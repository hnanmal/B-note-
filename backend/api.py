from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Response, Form
from fastapi.responses import StreamingResponse
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker
from typing import List, Optional
import pandas as pd
import io
import json
import datetime
import sqlite3
import ast
import operator

from . import crud, project_db, schemas, models
from .database import SessionLocal
from . import database


# 데이터베이스 세션을 가져오는 의존성
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _coerce_str(value) -> Optional[str]:
    if value is None:
        return None
    text_value = str(value).strip()
    return text_value or None


def _payload_get(obj, *keys):
    if not isinstance(obj, dict):
        return None
    for key in keys:
        if key in obj:
            value = obj.get(key)
            if value is not None and value != "":
                return value
    return None


def _sanitize_filename_part(value: str) -> str:
    raw = (value or "").strip() or "project"
    return (
        raw.replace("\\", "_")
        .replace("/", "_")
        .replace(":", "_")
        .replace("*", "_")
        .replace("?", "_")
        .replace('"', "_")
        .replace("<", "_")
        .replace(">", "_")
        .replace("|", "_")
    )


def _wm_trim(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = str(value).strip()
    if not cleaned or cleaned == "-":
        return None
    return cleaned


def _compose_spec_from_work_master_row(row: dict) -> Optional[str]:
    parts = [
        _wm_trim(row.get("cat_large_desc")),
        _wm_trim(row.get("cat_mid_desc")),
        _wm_trim(row.get("cat_small_desc")),
        _wm_trim(row.get("attr1_spec")),
        _wm_trim(row.get("attr2_spec")),
        _wm_trim(row.get("attr3_spec")),
    ]
    parts = [p for p in parts if p]
    return " | ".join(parts) if parts else None


_ALLOWED_BINOPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.FloorDiv: operator.floordiv,
    ast.Mod: operator.mod,
    ast.Pow: operator.pow,
}
_ALLOWED_UNARYOPS = {ast.UAdd: operator.pos, ast.USub: operator.neg}


def _try_parse_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except Exception:
        return None


def _safe_eval_numeric_expr(expr: str, variables: dict) -> Optional[float]:
    """Safely evaluate a numeric expression using only arithmetic + variables.

    Supported:
      - numbers (int/float)
      - variables: NAME
      - operators: + - * / // % ** and unary + -
      - parentheses
    """

    expr = (expr or "").strip()
    if expr.startswith("="):
        expr = expr[1:].strip()
    if not expr:
        return None

    direct = _try_parse_float(expr)
    if direct is not None:
        return direct

    try:
        tree = ast.parse(expr, mode="eval")
    except Exception:
        return None

    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)

        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return float(node.value)
            return None

        if isinstance(node, ast.Num):  # py<3.8
            return float(node.n)

        if isinstance(node, ast.Name):
            name = node.id
            if name not in variables:
                return None
            val = variables.get(name)
            return _try_parse_float(val)

        if isinstance(node, ast.UnaryOp):
            op = _ALLOWED_UNARYOPS.get(type(node.op))
            if not op:
                return None
            v = _eval(node.operand)
            if v is None:
                return None
            return float(op(v))

        if isinstance(node, ast.BinOp):
            op = _ALLOWED_BINOPS.get(type(node.op))
            if not op:
                return None
            left = _eval(node.left)
            right = _eval(node.right)
            if left is None or right is None:
                return None
            try:
                return float(op(left, right))
            except Exception:
                return None

        return None

    result = _eval(tree)
    if result is None:
        return None
    if isinstance(result, float) and (result != result):
        return None
    return float(result)


def get_project_db_session(project_identifier: str):
    try:
        db_path = project_db.resolve_project_db_path(project_identifier)
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    project_db.ensure_extra_tables(db_path)
    project_engine = create_engine(
        f"sqlite:///{db_path.as_posix()}",
        connect_args={"check_same_thread": False},
    )
    ProjectSessionLocal = sessionmaker(
        autocommit=False, autoflush=False, bind=project_engine
    )
    db = ProjectSessionLocal()
    try:
        yield db
    finally:
        db.close()
        project_engine.dispose()


router = APIRouter()


@router.get("/debug/ping", tags=["Debug"])
def debug_ping():
    return {"ok": True}


@router.get("/debug/db", tags=["Debug"])
def debug_db(db: Session = Depends(get_db)):
    try:
        cnt = db.query(models.StandardItem).count()
        return {"db_url": database.SQLALCHEMY_DATABASE_URL, "standard_items": cnt}
    except Exception as e:
        return {"db_url": database.SQLALCHEMY_DATABASE_URL, "error": str(e)}


# TODO: 실제 사용자 인증 로직으로 교체해야 합니다.
def get_current_user():
    # 임시로 고정된 사용자 ID를 반환합니다.
    return models.User(id=1, email="test@example.com", username="testuser")


@router.post("/projects/", response_model=schemas.Project)
def create_project_for_user(
    project: schemas.ProjectCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return crud.create_user_project(db=db, project=project, user_id=current_user.id)


# ===================
#        User
# ===================
@router.post("/users/", response_model=schemas.User, tags=["Users"])
def create_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    db_user = crud.get_user_by_email(db, email=user.email)
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    return crud.create_user(db=db, user=user)


# ===================
#     WorkMaster
# ===================
@router.post("/work-masters/", response_model=schemas.WorkMaster, tags=["Work Masters"])
def create_work_master(
    work_master: schemas.WorkMasterCreate, db: Session = Depends(get_db)
):
    db_work_master = crud.get_work_master_by_work_master_code(
        db, code=work_master.work_master_code
    )
    if db_work_master:
        raise HTTPException(
            status_code=400,
            detail=f"WorkMaster with code '{work_master.work_master_code}' already exists",
        )
    return crud.create_work_master(db=db, work_master=work_master)


@router.get(
    "/work-masters/", response_model=List[schemas.WorkMaster], tags=["Work Masters"]
)
def read_work_masters(
    skip: int = 0,
    limit: int = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
):
    work_masters = crud.get_work_masters(db, skip=skip, limit=limit, search=search)
    return work_masters


@router.get(
    "/work-masters/{work_master_id}",
    response_model=schemas.WorkMaster,
    tags=["Work Masters"],
)
def read_work_master(
    work_master_id: int,
    db: Session = Depends(get_db),
):
    db_work_master = crud.get_work_master(db, work_master_id=work_master_id)
    if not db_work_master:
        raise HTTPException(status_code=404, detail="WorkMaster not found")
    return db_work_master


@router.post(
    "/work-masters/upload",
    summary="Upload and upsert Work Masters from Excel",
    tags=["Work Masters"],
)
async def upload_work_masters(
    file: UploadFile = File(...), db: Session = Depends(get_db)
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Invalid file type. Please upload an .xlsx file."
        )

    created_count = 0
    updated_count = 0

    try:
        contents = await file.read()
        # 1. 4번째 행(header=3)을 컬럼명으로 읽고, 모든 데이터를 문자열(str)로 강제 변환합니다.
        df = pd.read_excel(io.BytesIO(contents), header=3, dtype=str)

        # 2. Pydantic 모델에 정의된 필드 이름을 순서대로 가져옵니다.
        model_fields = list(schemas.WorkMasterCreate.model_fields.keys())

        # 3. 엑셀 컬럼 개수와 모델 필드 개수가 다를 경우를 대비해, 모델 필드 개수만큼만 사용합니다.
        num_columns_to_use = min(len(df.columns), len(model_fields))
        df = df.iloc[:, :num_columns_to_use]
        df.columns = model_fields[:num_columns_to_use]

        for _, row in df.iterrows():
            # NaN 값을 None으로 변환하여 Pydantic 유효성 검사 통과
            row_data = row.where(pd.notna(row), None).to_dict()
            work_master_in = schemas.WorkMasterCreate(
                **row_data
            )  # 순서대로 매핑된 데이터로 객체 생성
            db_work_master = crud.get_work_master_by_work_master_code(
                db, code=work_master_in.work_master_code
            )

            if db_work_master:
                crud.update_work_master(
                    db, db_work_master=db_work_master, work_master_in=work_master_in
                )
                updated_count += 1
            else:
                crud.create_work_master(db, work_master=work_master_in)
                created_count += 1

        return {
            "message": "Work Masters uploaded successfully",
            "created": created_count,
            "updated": updated_count,
        }
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"An error occurred while processing the file: {e}"
        )


@router.patch(
    "/work-masters/{work_master_id}",
    response_model=schemas.WorkMaster,
    tags=["Work Masters"],
)
def update_work_master(
    work_master_id: int,
    updates: schemas.WorkMasterUpdate,
    db: Session = Depends(get_db),
):
    db_work_master = crud.get_work_master(db, work_master_id=work_master_id)
    if not db_work_master:
        raise HTTPException(status_code=404, detail="WorkMaster not found")
    payload = updates.dict(exclude_unset=True)
    if not payload:
        return db_work_master
    return crud.update_work_master_fields(
        db, db_work_master=db_work_master, updates=payload
    )


@router.get(
    "/project/{project_identifier}/work-masters/",
    response_model=List[schemas.WorkMaster],
    tags=["Project Data"],
)
def read_project_work_masters(
    project_identifier: str,
    skip: int = 0,
    limit: int = None,
    search: Optional[str] = None,
    db: Session = Depends(get_project_db_session),
):
    return crud.get_work_masters(db, skip=skip, limit=limit, search=search)


@router.post(
    "/project/{project_identifier}/work-masters/",
    response_model=schemas.WorkMaster,
    tags=["Project Data"],
)
def create_project_work_master(
    project_identifier: str,
    work_master: schemas.WorkMasterCreate,
    db: Session = Depends(get_project_db_session),
):
    db_work_master = crud.get_work_master_by_work_master_code(
        db, code=work_master.work_master_code
    )
    if db_work_master:
        raise HTTPException(
            status_code=400,
            detail=f"WorkMaster with code '{work_master.work_master_code}' already exists",
        )
    return crud.create_work_master(db=db, work_master=work_master)


# ===================
#  WorkMaster Precheck
# ===================
@router.get(
    "/project/{project_identifier}/work-masters/precheck",
    response_model=List[schemas.WorkMasterPrecheckState],
    tags=["Project Data"],
)
def read_project_work_master_precheck_states(
    project_identifier: str,
    db: Session = Depends(get_project_db_session),
):
    rows = db.execute(
        text(
            "SELECT work_master_id, use_yn, updated_at, other_opinion FROM work_master_precheck ORDER BY work_master_id"
        )
    ).fetchall()
    result: List[schemas.WorkMasterPrecheckState] = []
    for row in rows:
        result.append(
            schemas.WorkMasterPrecheckState(
                work_master_id=int(row[0]),
                use_yn=bool(row[1]),
                updated_at=row[2],
                other_opinion=row[3],
            )
        )
    return result


@router.patch(
    "/project/{project_identifier}/work-masters/{work_master_id}/precheck",
    response_model=schemas.WorkMasterPrecheckState,
    tags=["Project Data"],
)
def update_project_work_master_precheck_state(
    project_identifier: str,
    work_master_id: int,
    updates: schemas.WorkMasterPrecheckUpdate,
    db: Session = Depends(get_project_db_session),
):
    db_work_master = crud.get_work_master(db, work_master_id=work_master_id)
    if not db_work_master:
        raise HTTPException(status_code=404, detail="WorkMaster not found")

    if updates.use_yn is None and updates.other_opinion is None:
        raise HTTPException(status_code=400, detail="No updates provided")

    current = db.execute(
        text(
            "SELECT use_yn, other_opinion FROM work_master_precheck WHERE work_master_id = :work_master_id"
        ),
        {"work_master_id": work_master_id},
    ).fetchone()
    current_use = None
    current_other = None
    if current is not None:
        try:
            current_use = bool(current[0])
        except Exception:
            current_use = None
        current_other = current[1]

    now = datetime.datetime.utcnow().isoformat()
    next_use_bool = (
        updates.use_yn
        if updates.use_yn is not None
        else (current_use if current_use is not None else False)
    )
    use_value = 1 if next_use_bool else 0
    next_other_opinion = (
        updates.other_opinion if updates.other_opinion is not None else current_other
    )

    db.execute(
        text(
            """
            INSERT INTO work_master_precheck (work_master_id, use_yn, other_opinion, updated_at)
            VALUES (:work_master_id, :use_yn, :other_opinion, :updated_at)
            ON CONFLICT(work_master_id)
            DO UPDATE SET
              use_yn = excluded.use_yn,
              other_opinion = excluded.other_opinion,
              updated_at = excluded.updated_at
            """
        ),
        {
            "work_master_id": work_master_id,
            "use_yn": use_value,
            "other_opinion": next_other_opinion,
            "updated_at": now,
        },
    )
    db.commit()
    return schemas.WorkMasterPrecheckState(
        work_master_id=work_master_id,
        use_yn=bool(next_use_bool),
        updated_at=now,
        other_opinion=next_other_opinion,
    )


@router.post(
    "/project/{project_identifier}/work-masters/upload",
    summary="Upload and upsert Work Masters from Excel",
    tags=["Project Data"],
)
async def upload_project_work_masters(
    project_identifier: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_project_db_session),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Invalid file type. Please upload an .xlsx file."
        )

    created_count = 0
    updated_count = 0

    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents), header=3, dtype=str)
        model_fields = list(schemas.WorkMasterCreate.model_fields.keys())
        num_columns_to_use = min(len(df.columns), len(model_fields))
        df = df.iloc[:, :num_columns_to_use]
        df.columns = model_fields[:num_columns_to_use]

        for _, row in df.iterrows():
            row_data = row.where(pd.notna(row), None).to_dict()
            work_master_in = schemas.WorkMasterCreate(**row_data)
            db_work_master = crud.get_work_master_by_work_master_code(
                db, code=work_master_in.work_master_code
            )

            if db_work_master:
                crud.update_work_master(
                    db, db_work_master=db_work_master, work_master_in=work_master_in
                )
                updated_count += 1
            else:
                crud.create_work_master(db, work_master=work_master_in)
                created_count += 1

        return {
            "message": "Work Masters uploaded successfully",
            "created": created_count,
            "updated": updated_count,
        }
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"An error occurred while processing the file: {e}"
        )


def import_project_report_wm_excel_bytes(
    project_identifier: str,
    xlsx_bytes: bytes,
    db: Session,
):
    from openpyxl import load_workbook

    def norm(value) -> str:
        return str(value).strip() if value is not None else ""

    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid .xlsx file: {e}")

    if "Report_WM" not in wb.sheetnames:
        raise HTTPException(
            status_code=400, detail="Excel에 'Report_WM' 시트가 없습니다."
        )

    ws = wb["Report_WM"]
    max_col = int(ws.max_column or 0)
    if max_col <= 0:
        raise HTTPException(status_code=400, detail="Report_WM 시트가 비어 있습니다.")

    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    header_to_col = {norm(h): idx + 1 for idx, h in enumerate(headers) if norm(h)}

    col_wm_code = header_to_col.get("WM Code")
    col_gauge = header_to_col.get("Gauge")
    col_spec = header_to_col.get("Spec")
    col_other = header_to_col.get("기타의견")

    if not col_wm_code:
        raise HTTPException(
            status_code=400, detail="Report_WM 시트에 'WM Code' 컬럼이 없습니다."
        )

    if not col_spec and not col_other:
        raise HTTPException(
            status_code=400,
            detail="Report_WM 시트에 업데이트할 컬럼('Spec' 또는 '기타의견')이 없습니다.",
        )

    processed = 0
    matched = 0
    updated_spec = 0
    updated_other = 0
    missing = []

    max_row = int(ws.max_row or 1)
    for r in range(2, max_row + 1):
        wm_code = norm(ws.cell(row=r, column=col_wm_code).value)
        if not wm_code:
            continue

        gauge_norm = ""
        if col_gauge:
            gauge_norm = norm(ws.cell(row=r, column=col_gauge).value).upper()

        spec_value = None
        if col_spec:
            v = ws.cell(row=r, column=col_spec).value
            spec_value = "" if v is None else str(v)

        other_value = None
        if col_other:
            v = ws.cell(row=r, column=col_other).value
            other_value = "" if v is None else str(v)

        processed += 1

        row = db.execute(
            text(
                """
                SELECT id
                FROM work_masters
                WHERE work_master_code = :code
                  AND UPPER(COALESCE(TRIM(gauge), '')) = :gauge
                ORDER BY id
                LIMIT 1
                """
            ),
            {"code": wm_code, "gauge": gauge_norm},
        ).fetchone()

        work_master_id = int(row[0]) if row else None

        # Fallback: if gauge-based match fails, match by code only when unique.
        if work_master_id is None:
            rows = db.execute(
                text(
                    "SELECT id FROM work_masters WHERE work_master_code = :code ORDER BY id"
                ),
                {"code": wm_code},
            ).fetchall()
            if rows and len(rows) == 1:
                work_master_id = int(rows[0][0])

        if work_master_id is None:
            if len(missing) < 50:
                missing.append({"wm_code": wm_code, "gauge": gauge_norm})
            continue

        matched += 1
        now = datetime.datetime.utcnow().isoformat()

        if col_spec is not None and spec_value is not None:
            db.execute(
                text("UPDATE work_masters SET add_spec = :spec WHERE id = :id"),
                {"spec": spec_value, "id": work_master_id},
            )
            updated_spec += 1

        if col_other is not None and other_value is not None:
            current = db.execute(
                text(
                    "SELECT use_yn FROM work_master_precheck WHERE work_master_id = :id"
                ),
                {"id": work_master_id},
            ).fetchone()
            use_val = 1 if (current is not None and bool(current[0])) else 0
            db.execute(
                text(
                    """
                    INSERT INTO work_master_precheck (work_master_id, use_yn, other_opinion, updated_at)
                    VALUES (:id, :use_yn, :other, :updated_at)
                    ON CONFLICT(work_master_id)
                    DO UPDATE SET
                      other_opinion = excluded.other_opinion,
                      updated_at = excluded.updated_at
                    """
                ),
                {
                    "id": work_master_id,
                    "use_yn": use_val,
                    "other": other_value,
                    "updated_at": now,
                },
            )
            updated_other += 1

    db.commit()
    return {
        "processed_rows": processed,
        "matched_rows": matched,
        "updated_spec": updated_spec,
        "updated_other_opinion": updated_other,
        "missing": missing,
    }


@router.post(
    "/project/{project_identifier}/import/report-wm",
    summary="Import Report_WM sheet and update Spec/Other opinion",
    tags=["Project Data"],
)
async def import_project_report_wm_excel(
    project_identifier: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_project_db_session),
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Invalid file type. Please upload an .xlsx file."
        )
    contents = await file.read()
    return import_project_report_wm_excel_bytes(project_identifier, contents, db)


@router.patch(
    "/project/{project_identifier}/work-masters/{work_master_id}",
    response_model=schemas.WorkMaster,
    tags=["Project Data"],
)
def update_project_work_master(
    project_identifier: str,
    work_master_id: int,
    updates: schemas.WorkMasterUpdate,
    db: Session = Depends(get_project_db_session),
):
    db_work_master = crud.get_work_master(db, work_master_id=work_master_id)
    if not db_work_master:
        raise HTTPException(status_code=404, detail="WorkMaster not found")
    payload = updates.dict(exclude_unset=True)
    if not payload:
        return db_work_master
    return crud.update_work_master_fields(
        db, db_work_master=db_work_master, updates=payload
    )


@router.get(
    "/project/{project_identifier}/work-masters/{work_master_id}",
    response_model=schemas.WorkMaster,
    tags=["Project Data"],
)
def read_project_work_master(
    project_identifier: str,
    work_master_id: int,
    db: Session = Depends(get_project_db_session),
):
    db_work_master = crud.get_work_master(db, work_master_id=work_master_id)
    if not db_work_master:
        raise HTTPException(status_code=404, detail="WorkMaster not found")
    return db_work_master


@router.post(
    "/project/{project_identifier}/work-masters/{work_master_id}/add-gauge",
    response_model=schemas.WorkMaster,
    tags=["Project Data"],
)
def add_project_work_master_gauge(
    project_identifier: str,
    work_master_id: int,
    db: Session = Depends(get_project_db_session),
):
    db_work_master = crud.get_work_master(db, work_master_id=work_master_id)
    if not db_work_master:
        raise HTTPException(status_code=404, detail="WorkMaster not found")
    try:
        new_work_master = crud.duplicate_work_master_with_gauge(
            db, work_master_id=work_master_id
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not new_work_master:
        raise HTTPException(status_code=500, detail="게이지 항목을 생성할 수 없습니다.")
    return new_work_master


@router.post(
    "/project/{project_identifier}/work-masters/{work_master_id}/remove-gauge",
    tags=["Project Data"],
)
def remove_project_work_master_gauge(
    project_identifier: str,
    work_master_id: int,
    db: Session = Depends(get_project_db_session),
):
    db_work_master = crud.get_work_master(db, work_master_id=work_master_id)
    if not db_work_master:
        raise HTTPException(status_code=404, detail="WorkMaster not found")
    try:
        remaining = crud.remove_work_master_gauge(db, work_master_id=work_master_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if remaining is None:
        raise HTTPException(status_code=404, detail="게이지 항목을 삭제할 수 없습니다.")
    return {"remaining_gauges": len(remaining)}


# ===================
#  WorkMaster Cart
# ===================
def _normalize_cart_payload(raw_payload):
    def _ensure_list(value):
        return value if isinstance(value, list) else []

    revit_types = raw_payload.get("revitTypes") or raw_payload.get("revit_types") or []
    assignment_ids = (
        raw_payload.get("assignmentIds") or raw_payload.get("assignment_ids") or []
    )
    standard_item_ids = (
        raw_payload.get("standardItemIds") or raw_payload.get("standard_item_ids") or []
    )
    building_names = (
        raw_payload.get("buildingNames") or raw_payload.get("building_names") or []
    )
    formula = raw_payload.get("formula")
    return {
        "revit_types": _ensure_list(revit_types),
        "assignment_ids": _ensure_list(assignment_ids),
        "standard_item_ids": _ensure_list(standard_item_ids),
        "building_names": _ensure_list(building_names),
        "formula": formula,
    }


@router.get(
    "/project/{project_identifier}/workmaster-cart",
    response_model=List[schemas.WorkMasterCartEntry],
    tags=["Project Data"],
)
def read_project_workmaster_cart(
    project_identifier: str,
    db: Session = Depends(get_project_db_session),
):
    rows = db.execute(
        text(
            "SELECT id, payload, created_at FROM workmaster_cart_entries ORDER BY id DESC"
        )
    ).fetchall()
    entries: List[schemas.WorkMasterCartEntry] = []
    for row in rows:
        try:
            payload = json.loads(row[1] or "{}")
        except json.JSONDecodeError:
            payload = {}
        normalized = _normalize_cart_payload(payload)
        created_raw = row[2]
        try:
            created_at = (
                datetime.datetime.fromisoformat(created_raw)
                if created_raw
                else datetime.datetime.utcnow()
            )
        except ValueError:
            created_at = datetime.datetime.utcnow()
        entries.append(
            schemas.WorkMasterCartEntry(
                id=row[0],
                created_at=created_at,
                **normalized,
            )
        )
    return entries


@router.post(
    "/project/{project_identifier}/workmaster-cart",
    response_model=schemas.WorkMasterCartEntry,
    tags=["Project Data"],
)
def create_project_workmaster_cart_entry(
    project_identifier: str,
    payload: schemas.WorkMasterCartEntryCreate,
    db: Session = Depends(get_project_db_session),
):
    normalized = _normalize_cart_payload(payload.model_dump())
    now_iso = datetime.datetime.utcnow().isoformat()
    db.execute(
        text(
            "INSERT INTO workmaster_cart_entries (payload, created_at) VALUES (:payload, :created_at)"
        ),
        {"payload": json.dumps(normalized, ensure_ascii=False), "created_at": now_iso},
    )
    db.commit()
    new_id = db.execute(text("SELECT last_insert_rowid()")).scalar()
    created_at = datetime.datetime.fromisoformat(now_iso)
    return schemas.WorkMasterCartEntry(id=new_id, created_at=created_at, **normalized)


@router.patch(
    "/project/{project_identifier}/workmaster-cart/{entry_id}",
    response_model=schemas.WorkMasterCartEntry,
    tags=["Project Data"],
)
def update_project_workmaster_cart_entry(
    project_identifier: str,
    entry_id: int,
    payload: schemas.WorkMasterCartEntryUpdate,
    db: Session = Depends(get_project_db_session),
):
    row = db.execute(
        text(
            "SELECT payload, created_at FROM workmaster_cart_entries WHERE id = :entry_id"
        ),
        {"entry_id": entry_id},
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Cart entry not found")
    try:
        current_payload = json.loads(row[0] or "{}")
    except json.JSONDecodeError:
        current_payload = {}
    if payload.formula is not None:
        current_payload["formula"] = payload.formula
    normalized = _normalize_cart_payload(current_payload)
    db.execute(
        text(
            "UPDATE workmaster_cart_entries SET payload = :payload WHERE id = :entry_id"
        ),
        {"payload": json.dumps(normalized, ensure_ascii=False), "entry_id": entry_id},
    )
    db.commit()
    created_raw = row[1]
    try:
        created_at = (
            datetime.datetime.fromisoformat(created_raw)
            if created_raw
            else datetime.datetime.utcnow()
        )
    except ValueError:
        created_at = datetime.datetime.utcnow()
    return schemas.WorkMasterCartEntry(id=entry_id, created_at=created_at, **normalized)


@router.delete(
    "/project/{project_identifier}/workmaster-cart/{entry_id}",
    tags=["Project Data"],
)
def delete_project_workmaster_cart_entry(
    project_identifier: str,
    entry_id: int,
    db: Session = Depends(get_project_db_session),
):
    result = db.execute(
        text("DELETE FROM workmaster_cart_entries WHERE id = :entry_id"),
        {"entry_id": entry_id},
    )
    db.commit()
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Cart entry not found")
    return {"ok": True}


# ===================
#   StandardItem
# ===================
@router.get(
    "/standard-items/",
    response_model=List[schemas.StandardItem],
    tags=["Standard Items"],
)
def read_standard_items(
    skip: int = 0,
    limit: int = None,
    search: Optional[str] = None,
    parent_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    return crud.get_standard_items(
        db=db, skip=skip, limit=limit, search=search, parent_id=parent_id
    )


@router.post(
    "/standard-items/{standard_item_id}/assign",
    tags=["Standard Items"],
)
def assign_work_master(
    standard_item_id: int,
    payload: schemas.AssignWorkMaster,
    db: Session = Depends(get_db),
):
    std = crud.assign_work_master_to_standard_item(
        db, standard_item_id=standard_item_id, work_master_id=payload.work_master_id
    )
    if not std:
        raise HTTPException(
            status_code=404, detail="StandardItem or WorkMaster not found"
        )
    return {"message": "assigned", "standard_item_id": std.id}


@router.post(
    "/standard-items/{standard_item_id}/remove",
    tags=["Standard Items"],
)
def remove_work_master(
    standard_item_id: int,
    payload: schemas.AssignWorkMaster,
    db: Session = Depends(get_db),
):
    std = crud.remove_work_master_from_standard_item(
        db, standard_item_id=standard_item_id, work_master_id=payload.work_master_id
    )
    if not std:
        raise HTTPException(
            status_code=404, detail="StandardItem or WorkMaster not found"
        )
    return {"message": "removed", "standard_item_id": std.id}


@router.post(
    "/standard-items/{standard_item_id}/select",
    tags=["Standard Items"],
)
def select_standard_item_work_master(
    standard_item_id: int,
    selection: schemas.StandardItemWorkMasterSelectionRequest,
    db: Session = Depends(get_db),
):
    std = crud.get_standard_item(db, standard_item_id)
    if not std:
        raise HTTPException(status_code=404, detail="StandardItem not found")
    result = crud.select_work_master_for_standard_item(
        db, standard_item_id=standard_item_id, work_master_id=selection.work_master_id
    )
    return {"selected_work_master_id": result.work_master_id if result else None}


# Create standard item
@router.post(
    "/standard-items/", response_model=schemas.StandardItem, tags=["Standard Items"]
)
def create_standard_item(
    item: schemas.StandardItemCreate, db: Session = Depends(get_db)
):
    return crud.create_standard_item(db=db, standard_item=item)


@router.get(
    "/project/{project_identifier}/standard-items/",
    response_model=List[schemas.StandardItem],
    tags=["Project Data"],
)
def read_project_standard_items(
    project_identifier: str,
    skip: int = 0,
    limit: int = None,
    search: Optional[str] = None,
    parent_id: Optional[int] = None,
    db: Session = Depends(get_project_db_session),
):
    return crud.get_standard_items(
        db=db, skip=skip, limit=limit, search=search, parent_id=parent_id
    )


@router.get(
    "/project/{project_identifier}/standard-items/{standard_item_id}",
    response_model=schemas.StandardItem,
    tags=["Project Data"],
)
def get_project_standard_item(
    project_identifier: str,
    standard_item_id: int,
    db: Session = Depends(get_project_db_session),
):
    std = crud.get_standard_item(db, standard_item_id=standard_item_id)
    if not std:
        raise HTTPException(status_code=404, detail="StandardItem not found")
    return std


@router.post(
    "/project/{project_identifier}/standard-items/{standard_item_id}/assign",
    tags=["Project Data"],
)
def assign_work_master_to_project(
    project_identifier: str,
    standard_item_id: int,
    payload: schemas.AssignWorkMaster,
    db: Session = Depends(get_project_db_session),
):
    std = crud.assign_work_master_to_standard_item(
        db, standard_item_id=standard_item_id, work_master_id=payload.work_master_id
    )
    if not std:
        raise HTTPException(
            status_code=404, detail="StandardItem or WorkMaster not found"
        )
    return {"message": "assigned", "standard_item_id": std.id}


@router.post(
    "/project/{project_identifier}/standard-items/{standard_item_id}/remove",
    tags=["Project Data"],
)
def remove_project_work_master(
    project_identifier: str,
    standard_item_id: int,
    payload: schemas.AssignWorkMaster,
    db: Session = Depends(get_project_db_session),
):
    std = crud.remove_work_master_from_standard_item(
        db, standard_item_id=standard_item_id, work_master_id=payload.work_master_id
    )
    if not std:
        raise HTTPException(
            status_code=404, detail="StandardItem or WorkMaster not found"
        )
    return {"message": "removed", "standard_item_id": std.id}


@router.post(
    "/project/{project_identifier}/standard-items/{standard_item_id}/select",
    tags=["Project Data"],
)
def select_project_standard_item_work_master(
    project_identifier: str,
    standard_item_id: int,
    selection: schemas.StandardItemWorkMasterSelectionRequest,
    db: Session = Depends(get_project_db_session),
):
    std = crud.get_standard_item(db, standard_item_id)
    if not std:
        raise HTTPException(status_code=404, detail="StandardItem not found")
    result = crud.select_work_master_for_standard_item(
        db, standard_item_id=standard_item_id, work_master_id=selection.work_master_id
    )
    return {"selected_work_master_id": result.work_master_id if result else None}


@router.get(
    "/project/{project_identifier}/work-master-selections/summary",
    response_model=schemas.WorkMasterSummaryResponse,
    tags=["Project Data"],
)
def get_project_work_master_selection_summary(
    project_identifier: str,
    db: Session = Depends(get_project_db_session),
):
    rows = crud.list_selected_work_master_summary(db)
    return {"rows": rows}


@router.get(
    "/project/{project_identifier}/export/dynamo-json",
    response_model=schemas.DynamoProjectExportPayload,
    response_model_by_alias=True,
    tags=["Project Data"],
)
def export_project_db_for_dynamo(
    project_identifier: str,
    response: Response,
    db: Session = Depends(get_project_db_session),
):
    """Dynamo 테스트를 위한 프로젝트 DB JSON 추출 엔드포인트.

    - 추후에는 "파일 다운로드" 대신 Dynamo가 직접 참조하는 라우터로 사용 가능
    """

    buildings = crud.list_buildings(db)
    cart_entries = read_project_workmaster_cart(
        project_identifier=project_identifier, db=db
    )

    pjt_abbr = None
    try:
        pjt_abbr_row = db.execute(
            text("SELECT pjt_abbr FROM project_metadata ORDER BY id LIMIT 1")
        ).fetchone()
        if pjt_abbr_row and pjt_abbr_row[0]:
            pjt_abbr = str(pjt_abbr_row[0]).strip() or None
    except Exception:
        pjt_abbr = None

    try:
        now = datetime.datetime.now()
        stamp = now.strftime("%Y%m%d_%H%M%S")
        safe_abbr = (pjt_abbr or project_identifier or "project").strip()
        safe_abbr = (
            safe_abbr.replace("\\", "_")
            .replace("/", "_")
            .replace(":", "_")
            .replace("*", "_")
            .replace("?", "_")
            .replace('"', "_")
            .replace("<", "_")
            .replace(">", "_")
            .replace("|", "_")
        )
        filename = f"Bnote_{safe_abbr}_{stamp}.json"
        response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    except Exception:
        pass

    assignment_ids = sorted(
        {
            int(aid)
            for entry in cart_entries
            for aid in (getattr(entry, "assignment_ids", None) or [])
            if isinstance(aid, int) or (isinstance(aid, str) and aid.isdigit())
        }
    )
    standard_item_ids = sorted(
        {
            int(sid)
            for entry in cart_entries
            for sid in (getattr(entry, "standard_item_ids", None) or [])
            if isinstance(sid, int) or (isinstance(sid, str) and sid.isdigit())
        }
    )

    assignment_label_by_id = {}
    assignment_family_list_id_by_id = {}
    family_name_by_family_list_id = {}
    family_item_meta_by_id = {}
    assignment_standard_item_ids = set()
    assignment_family_name_by_id = {}
    assignment_standard_item_id_by_id = {}
    if assignment_ids:
        assigns = (
            db.query(models.GwmFamilyAssign)
            .filter(models.GwmFamilyAssign.id.in_(assignment_ids))
            .all()
        )
        for a in assigns:
            family_name = getattr(getattr(a, "family_list_item", None), "name", None)
            standard_item_id = getattr(a, "standard_item_id", None)
            if standard_item_id is not None:
                try:
                    assignment_standard_item_ids.add(int(standard_item_id))
                except (TypeError, ValueError):
                    pass
            if family_name:
                assignment_family_name_by_id[a.id] = family_name
            if standard_item_id is not None:
                try:
                    assignment_standard_item_id_by_id[a.id] = int(standard_item_id)
                except (TypeError, ValueError):
                    pass
            family_list_id = getattr(a, "family_list_id", None)
            if family_list_id is not None:
                try:
                    family_list_id_int = int(family_list_id)
                except (TypeError, ValueError):
                    family_list_id_int = None
                if family_list_id_int is not None:
                    assignment_family_list_id_by_id[a.id] = family_list_id_int
                    if family_name:
                        family_name_by_family_list_id[family_list_id_int] = family_name

    standard_item_name_by_id = {}
    standard_item_raw_name_by_id = {}
    standard_item_parent_id_by_id = {}
    standard_item_type_by_id = {}
    standard_item_ids_to_load = sorted(
        set(standard_item_ids) | set(assignment_standard_item_ids)
    )
    if standard_item_ids_to_load:
        loaded_standard_ids = set()
        pending_standard_ids = {int(sid) for sid in standard_item_ids_to_load}
        derive_from_by_id = {}
        while pending_standard_ids:
            batch_ids = sorted(pending_standard_ids - loaded_standard_ids)
            if not batch_ids:
                break
            rows = (
                db.query(
                    models.StandardItem.id,
                    models.StandardItem.name,
                    models.StandardItem.type,
                    models.StandardItem.parent_id,
                    models.StandardItem.derive_from,
                )
                .filter(models.StandardItem.id.in_(batch_ids))
                .all()
            )
            pending_standard_ids.clear()
            for sid, name, item_type, parent_id, derive_from in rows:
                sid_int = int(sid)
                loaded_standard_ids.add(sid_int)
                standard_item_name_by_id[sid_int] = name
                standard_item_raw_name_by_id[sid_int] = name
                standard_item_parent_id_by_id[sid_int] = (
                    int(parent_id) if parent_id is not None else None
                )
                standard_item_type_by_id[sid_int] = (
                    item_type.value if hasattr(item_type, "value") else str(item_type)
                )
                derive_from_by_id[sid_int] = derive_from
                if parent_id is not None:
                    try:
                        pending_standard_ids.add(int(parent_id))
                    except (TypeError, ValueError):
                        pass
                if derive_from is not None:
                    try:
                        pending_standard_ids.add(int(derive_from))
                    except (TypeError, ValueError):
                        pass

        formatted_name_by_id = {}
        for item_id, name in list(standard_item_name_by_id.items()):
            derive_from = derive_from_by_id.get(item_id)
            if derive_from is None:
                formatted_name_by_id[item_id] = name
                continue
            try:
                parent_id = int(derive_from)
            except (TypeError, ValueError):
                formatted_name_by_id[item_id] = name
                continue
            parent_name = standard_item_name_by_id.get(parent_id)
            if not parent_name:
                formatted_name_by_id[item_id] = name
                continue
            if pjt_abbr:
                formatted_name_by_id[item_id] = f"{parent_name} [{pjt_abbr}]::{name}"
            else:
                formatted_name_by_id[item_id] = f"{parent_name}::{name}"

        standard_item_name_by_id = formatted_name_by_id

    if assignment_ids:
        for aid in assignment_ids:
            family_name = assignment_family_name_by_id.get(aid)
            standard_item_id = assignment_standard_item_id_by_id.get(aid)
            standard_name = (
                standard_item_name_by_id.get(standard_item_id)
                if standard_item_id is not None
                else None
            )
            if family_name and standard_name:
                assignment_label_by_id[aid] = f"{family_name} / {standard_name}"
            elif standard_name:
                assignment_label_by_id[aid] = standard_name
            elif family_name:
                assignment_label_by_id[aid] = family_name
            else:
                assignment_label_by_id[aid] = f"assignment:{aid}"

    selected_work_master_by_standard_item_id = {}
    if standard_item_ids:
        rows = (
            db.query(
                models.StandardItemWorkMasterSelect.standard_item_id,
                models.WorkMaster.id,
                models.WorkMaster.work_master_code,
                models.WorkMaster.gauge,
                models.WorkMaster.discipline,
                models.WorkMaster.cat_large_desc,
                models.WorkMaster.cat_mid_desc,
                models.WorkMaster.cat_small_desc,
                models.WorkMaster.uom1,
                models.WorkMaster.uom2,
            )
            .join(
                models.WorkMaster,
                models.WorkMaster.id
                == models.StandardItemWorkMasterSelect.work_master_id,
            )
            .filter(
                models.StandardItemWorkMasterSelect.standard_item_id.in_(
                    standard_item_ids
                )
            )
            .all()
        )
        for (
            standard_item_id,
            work_master_id,
            work_master_code,
            gauge,
            discipline,
            cat_large_desc,
            cat_mid_desc,
            cat_small_desc,
            uom1,
            uom2,
        ) in rows:
            selected_work_master_by_standard_item_id[int(standard_item_id)] = {
                "id": int(work_master_id),
                "work_master_code": work_master_code,
                "gauge": gauge,
                "discipline": discipline,
                "cat_large_desc": cat_large_desc,
                "cat_mid_desc": cat_mid_desc,
                "cat_small_desc": cat_small_desc,
                "uom1": uom1,
                "uom2": uom2,
            }

    calc_dictionary_entries_by_family_list_id = {}
    family_list_ids = sorted({fid for fid in assignment_family_list_id_by_id.values()})

    def _load_family_items_with_ancestors(seed_ids):
        loaded_ids = set(family_item_meta_by_id.keys())
        pending_ids = {int(fid) for fid in (seed_ids or []) if fid}
        while pending_ids:
            batch = sorted(pending_ids - loaded_ids)
            if not batch:
                break
            rows = (
                db.query(
                    models.FamilyListItem.id,
                    models.FamilyListItem.parent_id,
                    models.FamilyListItem.name,
                    models.FamilyListItem.sequence_number,
                )
                .filter(models.FamilyListItem.id.in_(batch))
                .all()
            )
            pending_ids.clear()
            for fid, parent_id, name, sequence_number in rows:
                fid_int = int(fid)
                loaded_ids.add(fid_int)
                family_item_meta_by_id[fid_int] = {
                    "id": fid_int,
                    "parent_id": int(parent_id) if parent_id is not None else None,
                    "name": name,
                    "sequence_number": sequence_number,
                }
                if parent_id is not None:
                    try:
                        pending_ids.add(int(parent_id))
                    except (TypeError, ValueError):
                        pass

    _load_family_items_with_ancestors(family_list_ids)

    def _family_path(fid):
        if not fid:
            return []
        path = []
        cursor = int(fid)
        seen = set()
        while cursor and cursor not in seen:
            seen.add(cursor)
            meta = family_item_meta_by_id.get(cursor)
            if not meta:
                break
            path.append(meta)
            cursor = meta.get("parent_id")
        return list(reversed(path))

    def _category_label_from_family_root(root_meta):
        if not root_meta:
            return None
        seq = root_meta.get("sequence_number")
        name = root_meta.get("name")
        if seq and name:
            return f"{str(seq).strip()}.{str(name).strip()}"
        return str(name).strip() if name else None

    def _parse_standard_type_from_family(level2_meta):
        if not level2_meta:
            return (None, None)
        seq = level2_meta.get("sequence_number")
        name = level2_meta.get("name")
        if seq:
            seq = str(seq).strip() or None
        if name:
            name = str(name).strip() or None
        if seq or name:
            return (seq, name)

        combined = str(level2_meta.get("name") or "").strip()
        if not combined:
            return (None, None)
        import re

        m = re.match(r"^\s*([0-9]+(?:\.[0-9]+)*)\s+(.+?)\s*$", combined)
        if m:
            return (m.group(1), m.group(2))
        m = re.match(r"^\s*([0-9]+)\.(.+?)\s*$", combined)
        if m:
            return (m.group(1), m.group(2).strip())
        return (None, combined)

    assignment_category_by_id = {}
    assignment_family_std_type_number_by_id = {}
    assignment_family_std_type_name_by_id = {}
    if assignment_ids:
        for aid in assignment_ids:
            fid = assignment_family_list_id_by_id.get(aid)
            path = _family_path(fid)
            if not path:
                continue
            assignment_category_by_id[aid] = _category_label_from_family_root(path[0])
            level2_meta = path[2] if len(path) > 2 else path[-1]
            std_no, std_name = _parse_standard_type_from_family(level2_meta)
            assignment_family_std_type_number_by_id[aid] = std_no
            assignment_family_std_type_name_by_id[aid] = std_name
    if family_list_ids:
        calc_entries = (
            db.query(models.CalcDictionaryEntry)
            .filter(models.CalcDictionaryEntry.family_list_id.in_(family_list_ids))
            .order_by(
                models.CalcDictionaryEntry.family_list_id,
                models.CalcDictionaryEntry.symbol_key,
            )
            .all()
        )
        for entry in calc_entries:
            fid = int(getattr(entry, "family_list_id", 0) or 0)
            if not fid:
                continue
            calc_dictionary_entries_by_family_list_id.setdefault(fid, []).append(
                schemas.CalcDictionarySymbol(
                    family_list_id=fid,
                    family_name=family_name_by_family_list_id.get(fid),
                    calc_code=getattr(entry, "calc_code", None),
                    symbol_key=getattr(entry, "symbol_key", ""),
                    symbol_value=getattr(entry, "symbol_value", ""),
                )
            )

    for entry in cart_entries:
        entry.assignment_labels = [
            assignment_label_by_id.get(int(aid), f"assignment:{aid}")
            for aid in (entry.assignment_ids or [])
        ]
        entry.standard_item_names = [
            standard_item_name_by_id.get(int(sid), f"standard_item:{sid}")
            for sid in (entry.standard_item_ids or [])
        ]
        entry.work_masters = [
            schemas.WorkMasterBrief(**wm)
            for sid in (entry.standard_item_ids or [])
            for wm in [selected_work_master_by_standard_item_id.get(int(sid))]
            if wm
        ]

        family_list_ids_for_entry = []
        seen_family_list_ids = set()
        for aid in entry.assignment_ids or []:
            try:
                aid_int = int(aid)
            except (TypeError, ValueError):
                continue
            fid = assignment_family_list_id_by_id.get(aid_int)
            if not fid or fid in seen_family_list_ids:
                continue
            seen_family_list_ids.add(fid)
            family_list_ids_for_entry.append(fid)

        entry.calc_dictionary_entries = [
            calc_entry
            for fid in family_list_ids_for_entry
            for calc_entry in calc_dictionary_entries_by_family_list_id.get(fid, [])
        ]

    def _first_from(value):
        if value is None:
            return None
        if isinstance(value, (list, tuple)):
            return value[0] if value else None
        return value

    def _coerce_int(value):
        if value is None:
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            if value.isdigit():
                return int(value)
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _coerce_str(value):
        if value is None:
            return None
        text_value = str(value).strip()
        return text_value or None

    def _standard_tree_level2_name(standard_item_id):
        sid = _coerce_int(standard_item_id)
        if not sid:
            return None
        path_ids = []
        cursor = sid
        seen = set()
        while cursor and cursor not in seen:
            seen.add(cursor)
            path_ids.append(cursor)
            cursor = standard_item_parent_id_by_id.get(cursor)
        path_ids = list(reversed(path_ids))
        if not path_ids:
            return None
        level2_id = path_ids[2] if len(path_ids) > 2 else path_ids[-1]
        return _coerce_str(standard_item_raw_name_by_id.get(level2_id))

    def _standard_tree_level1_name(standard_item_id):
        sid = _coerce_int(standard_item_id)
        if not sid:
            return None
        path_ids = []
        cursor = sid
        seen = set()
        while cursor and cursor not in seen:
            seen.add(cursor)
            path_ids.append(cursor)
            cursor = standard_item_parent_id_by_id.get(cursor)
        path_ids = list(reversed(path_ids))
        if not path_ids:
            return None
        level1_id = path_ids[1] if len(path_ids) > 1 else path_ids[-1]
        return _coerce_str(standard_item_raw_name_by_id.get(level1_id))

    def _compose_unit(wm_obj):
        if wm_obj is None:
            return None
        return _coerce_str(getattr(wm_obj, "uom1", None))

    dynamo_cart_entries = []
    for entry in cart_entries:
        wm = _first_from(getattr(entry, "work_masters", None) or [])
        if wm is not None and not isinstance(wm, schemas.WorkMasterBrief):
            try:
                wm = schemas.WorkMasterBrief(**wm)
            except Exception:
                wm = None

        standard_item_id_value = _coerce_int(
            _first_from(getattr(entry, "standard_item_ids", None) or [])
        )
        standard_item_name_value = _coerce_str(
            _first_from(getattr(entry, "standard_item_names", None) or [])
        )

        assignment_id_value = _coerce_int(
            _first_from(getattr(entry, "assignment_ids", None) or [])
        )
        family_category_value = (
            assignment_category_by_id.get(assignment_id_value)
            if assignment_id_value is not None
            else None
        )
        family_std_type_number_value = (
            assignment_family_std_type_number_by_id.get(assignment_id_value)
            if assignment_id_value is not None
            else None
        )
        family_std_type_name_value = (
            assignment_family_std_type_name_by_id.get(assignment_id_value)
            if assignment_id_value is not None
            else None
        )
        standard_item_type_value = (
            _coerce_str(standard_item_type_by_id.get(standard_item_id_value))
            if standard_item_id_value is not None
            else None
        )

        standard_tree_level1_value = _standard_tree_level1_name(standard_item_id_value)
        standard_tree_level2_value = _standard_tree_level2_name(standard_item_id_value)
        if standard_tree_level1_value and standard_tree_level2_value:
            detail_classification_value = (
                f"{standard_tree_level1_value} | {standard_tree_level2_value}"
            )
        else:
            detail_classification_value = (
                standard_tree_level2_value
                or standard_tree_level1_value
                or standard_item_name_value
            )

        dynamo_cart_entries.append(
            schemas.DynamoWorkMasterCartEntry(
                id=int(getattr(entry, "id")),
                created_at=getattr(entry, "created_at"),
                formula=getattr(entry, "formula", None),
                revit_type=_coerce_str(
                    _first_from(getattr(entry, "revit_types", None) or [])
                ),
                assignment_id=assignment_id_value,
                standard_item_id=standard_item_id_value,
                building_name=_coerce_str(
                    _first_from(getattr(entry, "building_names", None) or [])
                ),
                assignment_label=_coerce_str(
                    _first_from(getattr(entry, "assignment_labels", None) or [])
                ),
                standard_item_name=standard_item_name_value,
                category=_coerce_str(family_category_value),
                standard_type_number=_coerce_str(family_std_type_number_value),
                standard_type_name=_coerce_str(family_std_type_name_value),
                classification=standard_item_type_value,
                detail_classification=detail_classification_value,
                unit=_compose_unit(wm),
                work_master=wm,
                calc_dictionary_entries=list(
                    getattr(entry, "calc_dictionary_entries", None) or []
                ),
            )
        )
    return {
        "project_identifier": project_identifier,
        "buildings": buildings,
        "workmaster_cart_entries": dynamo_cart_entries,
    }


@router.get(
    "/project/{project_identifier}/export/db-json",
    response_model=schemas.DynamoProjectExportPayload,
    response_model_by_alias=True,
    tags=["Project Data"],
)
def export_project_db_json(
    project_identifier: str,
    response: Response,
    db: Session = Depends(get_project_db_session),
):
    """Compatibility alias for the Dynamo JSON export.

    Frontend uses `/export/db-json` for the Dynamo download button.
    """

    return export_project_db_for_dynamo(
        project_identifier=project_identifier, response=response, db=db
    )


@router.post(
    "/project/{project_identifier}/calc-result/import-json",
    response_model=schemas.CalcResultImportResponse,
    tags=["Project Data"],
)
def import_calc_result_json(
    project_identifier: str,
    rev_key: str = Form(...),
    mode: str = Form("append"),
    file: UploadFile = File(...),
    db: Session = Depends(get_project_db_session),
):
    rev_key = _coerce_str(rev_key)
    if not rev_key:
        raise HTTPException(status_code=400, detail="rev_key is required")
    mode = (_coerce_str(mode) or "append").lower()
    if mode not in {"append", "overwrite"}:
        raise HTTPException(status_code=400, detail="mode must be append|overwrite")

    try:
        raw = file.file.read()
        payload = json.loads(raw.decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid JSON file")

    project_info = payload.get("project_info") if isinstance(payload, dict) else None
    building_name = _coerce_str(
        _payload_get(
            project_info or {},
            "building name",
            "building_name",
            "buildingName",
        )
    )
    results = None
    if isinstance(payload, dict):
        results = payload.get("calculation result")
        if results is None:
            results = payload.get("calculation_result")
        if results is None:
            results = payload.get("calculationResult")
    if results is None and isinstance(payload, list):
        results = payload
    if not isinstance(results, list):
        results = []

    if building_name:
        try:
            existing = (
                db.query(models.BuildingList)
                .filter(models.BuildingList.name == building_name)
                .first()
            )
            if not existing:
                db.add(models.BuildingList(name=building_name))
                db.commit()
        except Exception:
            db.rollback()

    deleted = 0
    if mode == "overwrite":
        if not building_name:
            raise HTTPException(
                status_code=400,
                detail="building name is required in JSON for overwrite mode",
            )
        try:
            res = db.execute(
                text(
                    "DELETE FROM calc_result WHERE TRIM(COALESCE(building_name,'')) = TRIM(:building_name) AND rev_key = :rev_key"
                ),
                {"building_name": building_name, "rev_key": rev_key},
            )
            deleted = int(getattr(res, "rowcount", 0) or 0)
            db.commit()
        except Exception:
            db.rollback()
            raise HTTPException(
                status_code=500, detail="Failed to overwrite calc results"
            )

    now_iso = datetime.datetime.utcnow().isoformat()

    def _resolve_work_master_id(
        work_master_id: Optional[int], work_master_code: Optional[str]
    ) -> Optional[int]:
        if work_master_id:
            return int(work_master_id)
        code = _coerce_str(work_master_code)
        if not code:
            return None
        try:
            row = db.execute(
                text(
                    "SELECT id FROM work_masters WHERE work_master_code = :code LIMIT 1"
                ),
                {"code": code},
            ).fetchone()
            return int(row[0]) if row and row[0] is not None else None
        except Exception:
            return None

    def _mk_key(*parts) -> str:
        return "|".join([_sanitize_filename_part(_coerce_str(p) or "") for p in parts])

    inserted = 0
    for entry in results:
        if not isinstance(entry, dict):
            continue
        wm_payload = (
            entry.get("work_master")
            if isinstance(entry.get("work_master"), dict)
            else {}
        )

        work_master_id = _payload_get(
            wm_payload, "id", "work_master_id", "workMasterId"
        )
        work_master_code = _coerce_str(
            _payload_get(
                wm_payload, "work_master_code", "workMasterCode", "wm_code", "wmCode"
            )
        )
        resolved_wm_id = _resolve_work_master_id(
            (
                int(work_master_id)
                if work_master_id is not None and str(work_master_id).isdigit()
                else None
            ),
            work_master_code,
        )

        guid = _coerce_str(_payload_get(entry, "GUID", "guid"))
        gui = _coerce_str(_payload_get(entry, "GUI", "gui"))
        member_name = _coerce_str(
            _payload_get(entry, "name", "member", "member_name", "memberName")
        )
        category = _coerce_str(_payload_get(entry, "카테고리", "category"))
        std_num = _coerce_str(
            _payload_get(
                entry, "표준타입 번호", "standard_type_number", "standardTypeNumber"
            )
        )
        std_name = _coerce_str(
            _payload_get(
                entry, "표준타입 이름", "standard_type_name", "standardTypeName"
            )
        )
        classification = _coerce_str(_payload_get(entry, "분류", "classification"))
        detail = _coerce_str(
            _payload_get(
                entry, "상세분류", "detail_classification", "detailClassification"
            )
        )
        unit = _coerce_str(_payload_get(entry, "단위", "unit"))
        formula = _coerce_str(_payload_get(entry, "수식", "formula"))
        substituted = _coerce_str(
            _payload_get(entry, "대입수식", "substituted_formula", "substitutedFormula")
        )
        result_val = _payload_get(entry, "산출결과", "result")
        result_log = _coerce_str(
            _payload_get(entry, "산출로그", "result_log", "resultLog")
        )

        result_float = None
        if isinstance(result_val, (int, float)):
            result_float = float(result_val)
        elif result_val is not None:
            try:
                result_float = float(str(result_val).strip())
            except Exception:
                result_float = None

        key = _mk_key(
            rev_key or "",
            building_name or "",
            guid or "",
            gui or "",
            formula or "",
            work_master_code or str(resolved_wm_id or ""),
            detail or "",
        )

        try:
            db.execute(
                text(
                    """
                    INSERT OR REPLACE INTO calc_result (
                        key, rev_key, building_name, guid, gui, member_name,
                        category, standard_type_number, standard_type_name,
                        classification, detail_classification, unit,
                        formula, substituted_formula, result, result_log,
                        work_master_id, work_master_code, created_at
                    ) VALUES (
                        :key, :rev_key, :building_name, :guid, :gui, :member_name,
                        :category, :standard_type_number, :standard_type_name,
                        :classification, :detail_classification, :unit,
                        :formula, :substituted_formula, :result, :result_log,
                        :work_master_id, :work_master_code, :created_at
                    )
                    """
                ),
                {
                    "key": key,
                    "rev_key": rev_key,
                    "building_name": building_name,
                    "guid": guid,
                    "gui": gui,
                    "member_name": member_name,
                    "category": category,
                    "standard_type_number": std_num,
                    "standard_type_name": std_name,
                    "classification": classification,
                    "detail_classification": detail,
                    "unit": unit,
                    "formula": formula,
                    "substituted_formula": substituted,
                    "result": result_float,
                    "result_log": result_log,
                    "work_master_id": resolved_wm_id,
                    "work_master_code": work_master_code,
                    "created_at": now_iso,
                },
            )
            inserted += 1
        except Exception:
            continue

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Failed to store calc results")

    return {
        "project_identifier": project_identifier,
        "building_name": building_name,
        "rev_key": rev_key,
        "mode": mode,
        "deleted": deleted,
        "inserted": inserted,
    }


@router.get(
    "/project/{project_identifier}/calc-result",
    response_model=List[schemas.CalcResultRow],
    tags=["Project Data"],
)
def list_calc_results(
    project_identifier: str,
    building_name: Optional[str] = None,
    rev_key: Optional[str] = None,
    limit: int = 2000,
    offset: int = 0,
    db: Session = Depends(get_project_db_session),
):
    limit = max(1, min(int(limit), 20000))
    offset = max(0, int(offset))
    building_name = _coerce_str(building_name)
    rev_key = _coerce_str(rev_key)

    rows = (
        db.execute(
            text(
                """
            SELECT
                cr.id AS id,
                cr.created_at AS created_at,
                cr.building_name AS building_name,
                cr.rev_key AS rev_key,
                cr.category AS category,
                cr.standard_type_number AS standard_type_number,
                cr.standard_type_name AS standard_type_name,
                cr.classification AS classification,
                cr.detail_classification AS detail_classification,
                cr.unit AS unit,
                cr.formula AS formula,
                cr.substituted_formula AS substituted_formula,
                cr.result AS result,
                cr.result_log AS result_log,
                cr.guid AS guid,
                cr.gui AS gui,
                cr.member_name AS member_name,
                COALESCE(wm.work_master_code, cr.work_master_code) AS wm_code,
                wm.gauge AS gauge,
                wm.add_spec AS add_spec,
                wm.cat_large_desc AS cat_large_desc,
                wm.cat_mid_desc AS cat_mid_desc,
                wm.cat_small_desc AS cat_small_desc,
                wm.attr1_spec AS attr1_spec,
                wm.attr2_spec AS attr2_spec,
                wm.attr3_spec AS attr3_spec
            FROM calc_result cr
            LEFT JOIN work_masters wm
              ON (wm.id = cr.work_master_id)
              OR (cr.work_master_id IS NULL AND wm.work_master_code = cr.work_master_code)
                        WHERE (:building_name IS NULL OR cr.building_name = :building_name)
                            AND (:rev_key IS NULL OR cr.rev_key = :rev_key)
            ORDER BY cr.id DESC
            LIMIT :limit OFFSET :offset
            """
            ),
            {
                "building_name": building_name,
                "rev_key": rev_key,
                "limit": limit,
                "offset": offset,
            },
        )
        .mappings()
        .all()
    )

    output: List[schemas.CalcResultRow] = []
    for row in rows:
        row_dict = dict(row)
        output.append(
            schemas.CalcResultRow(
                id=int(row_dict.get("id")),
                created_at=str(row_dict.get("created_at")),
                building_name=_coerce_str(row_dict.get("building_name")),
                rev_key=_coerce_str(row_dict.get("rev_key")),
                category=_coerce_str(row_dict.get("category")),
                standard_type_number=_coerce_str(row_dict.get("standard_type_number")),
                standard_type_name=_coerce_str(row_dict.get("standard_type_name")),
                classification=_coerce_str(row_dict.get("classification")),
                description=_coerce_str(row_dict.get("detail_classification")),
                guid=_coerce_str(row_dict.get("guid")),
                gui=_coerce_str(row_dict.get("gui")),
                member_name=_coerce_str(row_dict.get("member_name")),
                wm_code=_coerce_str(row_dict.get("wm_code")),
                gauge=_coerce_str(row_dict.get("gauge")),
                spec=_compose_spec_from_work_master_row(row_dict),
                add_spec=_coerce_str(row_dict.get("add_spec")),
                formula=_coerce_str(row_dict.get("formula")),
                substituted_formula=_coerce_str(row_dict.get("substituted_formula")),
                result=(
                    float(row_dict["result"])
                    if row_dict.get("result") is not None
                    else None
                ),
                result_log=_coerce_str(row_dict.get("result_log")),
                unit=_coerce_str(row_dict.get("unit")),
            )
        )
    return output


@router.get(
    "/project/{project_identifier}/calc-result/rev-keys",
    response_model=List[str],
    tags=["Project Data"],
)
def list_calc_result_rev_keys(
    project_identifier: str,
    building_name: Optional[str] = None,
    db: Session = Depends(get_project_db_session),
):
    building_name = _coerce_str(building_name)
    rows = db.execute(
        text(
            """
                SELECT DISTINCT rev_key
                FROM calc_result
                WHERE rev_key IS NOT NULL AND rev_key != ''
                                    AND (
                                        :building_name IS NULL
                                        OR TRIM(COALESCE(building_name, '')) = TRIM(:building_name)
                                    )
                ORDER BY rev_key
                """
        ),
        {"building_name": building_name},
    ).fetchall()
    return [str(r[0]) for r in rows if r and r[0] is not None]


@router.get(
    "/project/{project_identifier}/calc-result/buildings",
    response_model=List[str],
    tags=["Project Data"],
)
def list_calc_result_buildings(
    project_identifier: str,
    db: Session = Depends(get_project_db_session),
):
    rows = db.execute(
        text(
            """
                SELECT DISTINCT building_name
                FROM calc_result
                WHERE building_name IS NOT NULL AND building_name != ''
                ORDER BY building_name
                """
        )
    ).fetchall()
    return [str(r[0]) for r in rows if r and r[0] is not None]


@router.post(
    "/project/{project_identifier}/calc-result/manual-update",
    response_model=schemas.CalcResultManualUpdateResponse,
    tags=["Project Data"],
)
def manual_update_calc_results(
    project_identifier: str,
    rev_key: str = Form(...),
    db: Session = Depends(get_project_db_session),
):
    rev_key = _coerce_str(rev_key)
    if not rev_key:
        raise HTTPException(status_code=400, detail="rev_key is required")

    buildings = db.execute(
        text(
            "SELECT name FROM building_list WHERE name IS NOT NULL AND TRIM(name) != '' ORDER BY name"
        )
    ).fetchall()
    building_names = [str(r[0]) for r in buildings if r and r[0] is not None]
    if not building_names:
        raise HTTPException(status_code=400, detail="No buildings found")

    rows = db.execute(
        text("SELECT id, payload FROM workmaster_cart_entries ORDER BY id DESC")
    ).fetchall()

    cart_entries_scanned = 0
    inserted = 0
    skipped = 0
    manual_entries_matched = 0

    # Collect ids for family + selected work master lookups
    cart_entry_payloads = []
    assignment_ids = set()
    standard_item_ids = set()
    for row in rows:
        cart_entries_scanned += 1
        entry_id = int(row[0])
        try:
            payload = json.loads(row[1] or "{}")
        except Exception:
            payload = {}
        normalized = _normalize_cart_payload(payload)
        aids = normalized.get("assignment_ids") or []
        for aid in aids:
            try:
                assignment_ids.add(int(aid))
            except Exception:
                pass
        sids = normalized.get("standard_item_ids") or []
        for sid in sids:
            try:
                standard_item_ids.add(int(sid))
            except Exception:
                pass
        cart_entry_payloads.append((entry_id, normalized))

    assignment_family_list_id_by_id = {}
    family_item_meta_by_id = {}
    if assignment_ids:
        assigns = (
            db.query(models.GwmFamilyAssign)
            .filter(models.GwmFamilyAssign.id.in_(sorted(assignment_ids)))
            .all()
        )
        for a in assigns:
            fid = getattr(a, "family_list_id", None)
            if fid is None:
                continue
            try:
                assignment_family_list_id_by_id[int(a.id)] = int(fid)
            except Exception:
                continue

    family_list_ids = sorted(
        {fid for fid in assignment_family_list_id_by_id.values() if fid}
    )

    # Load selected WorkMaster (for wm_code/gauge/spec/add_spec/unit via list endpoint JOIN)
    selected_work_master_by_standard_item_id = {}
    if standard_item_ids:
        rows_wm = (
            db.query(
                models.StandardItemWorkMasterSelect.standard_item_id,
                models.WorkMaster.id,
                models.WorkMaster.work_master_code,
                models.WorkMaster.gauge,
                models.WorkMaster.add_spec,
                models.WorkMaster.uom1,
                models.WorkMaster.cat_large_desc,
                models.WorkMaster.cat_mid_desc,
                models.WorkMaster.cat_small_desc,
                models.WorkMaster.attr1_spec,
                models.WorkMaster.attr2_spec,
                models.WorkMaster.attr3_spec,
            )
            .join(
                models.WorkMaster,
                models.WorkMaster.id
                == models.StandardItemWorkMasterSelect.work_master_id,
            )
            .filter(
                models.StandardItemWorkMasterSelect.standard_item_id.in_(
                    sorted(standard_item_ids)
                )
            )
            .all()
        )
        for (
            sid,
            wm_id,
            wm_code,
            gauge,
            add_spec,
            uom1,
            cat_large_desc,
            cat_mid_desc,
            cat_small_desc,
            attr1_spec,
            attr2_spec,
            attr3_spec,
        ) in rows_wm:
            selected_work_master_by_standard_item_id[int(sid)] = {
                "work_master_id": int(wm_id) if wm_id is not None else None,
                "work_master_code": _coerce_str(wm_code),
                "gauge": _coerce_str(gauge),
                "add_spec": _coerce_str(add_spec),
                "uom1": _coerce_str(uom1),
                "cat_large_desc": _coerce_str(cat_large_desc),
                "cat_mid_desc": _coerce_str(cat_mid_desc),
                "cat_small_desc": _coerce_str(cat_small_desc),
                "attr1_spec": _coerce_str(attr1_spec),
                "attr2_spec": _coerce_str(attr2_spec),
                "attr3_spec": _coerce_str(attr3_spec),
            }

    # Load StandardItem tree info to build detail_classification like Dynamo export
    standard_item_name_by_id = {}
    standard_item_raw_name_by_id = {}
    standard_item_parent_id_by_id = {}
    standard_item_type_by_id = {}
    if standard_item_ids:
        loaded_ids = set()
        pending_ids = {int(sid) for sid in standard_item_ids if sid}
        while pending_ids:
            batch_ids = sorted(pending_ids - loaded_ids)
            if not batch_ids:
                break
            rows_std = (
                db.query(
                    models.StandardItem.id,
                    models.StandardItem.name,
                    models.StandardItem.type,
                    models.StandardItem.parent_id,
                )
                .filter(models.StandardItem.id.in_(batch_ids))
                .all()
            )
            pending_ids.clear()
            for sid, name, item_type, parent_id in rows_std:
                sid_int = int(sid)
                loaded_ids.add(sid_int)
                standard_item_name_by_id[sid_int] = name
                standard_item_raw_name_by_id[sid_int] = name
                standard_item_parent_id_by_id[sid_int] = (
                    int(parent_id) if parent_id is not None else None
                )
                standard_item_type_by_id[sid_int] = (
                    item_type.value if hasattr(item_type, "value") else str(item_type)
                )
                if parent_id is not None:
                    try:
                        pending_ids.add(int(parent_id))
                    except Exception:
                        pass

    def _standard_tree_level2_name(standard_item_id: Optional[int]) -> Optional[str]:
        try:
            sid = int(standard_item_id) if standard_item_id is not None else 0
        except Exception:
            sid = 0
        if not sid:
            return None
        path_ids = []
        cursor = sid
        seen = set()
        while cursor and cursor not in seen:
            seen.add(cursor)
            path_ids.append(cursor)
            cursor = standard_item_parent_id_by_id.get(cursor)
        path_ids = list(reversed(path_ids))
        if not path_ids:
            return None
        level2_id = path_ids[2] if len(path_ids) > 2 else path_ids[-1]
        return _coerce_str(standard_item_raw_name_by_id.get(level2_id))

    def _standard_tree_level1_name(standard_item_id: Optional[int]) -> Optional[str]:
        try:
            sid = int(standard_item_id) if standard_item_id is not None else 0
        except Exception:
            sid = 0
        if not sid:
            return None
        path_ids = []
        cursor = sid
        seen = set()
        while cursor and cursor not in seen:
            seen.add(cursor)
            path_ids.append(cursor)
            cursor = standard_item_parent_id_by_id.get(cursor)
        path_ids = list(reversed(path_ids))
        if not path_ids:
            return None
        level1_id = path_ids[1] if len(path_ids) > 1 else path_ids[-1]
        return _coerce_str(standard_item_raw_name_by_id.get(level1_id))

    def _compose_unit_from_wm_meta(wm_meta: Optional[dict]) -> Optional[str]:
        if not wm_meta:
            return None
        return _coerce_str(wm_meta.get("uom1"))

    def _compose_spec_from_wm_meta(wm_meta: Optional[dict]) -> Optional[str]:
        if not wm_meta:
            return None
        return _compose_spec_from_work_master_row(wm_meta)

    def _load_family_items_with_ancestors(seed_ids):
        loaded_ids = set(family_item_meta_by_id.keys())
        pending_ids = {int(fid) for fid in (seed_ids or []) if fid}
        while pending_ids:
            batch = sorted(pending_ids - loaded_ids)
            if not batch:
                break
            rows_ = (
                db.query(
                    models.FamilyListItem.id,
                    models.FamilyListItem.parent_id,
                    models.FamilyListItem.name,
                    models.FamilyListItem.sequence_number,
                )
                .filter(models.FamilyListItem.id.in_(batch))
                .all()
            )
            pending_ids.clear()
            for fid, parent_id, name, sequence_number in rows_:
                fid_int = int(fid)
                loaded_ids.add(fid_int)
                family_item_meta_by_id[fid_int] = {
                    "id": fid_int,
                    "parent_id": int(parent_id) if parent_id is not None else None,
                    "name": name,
                    "sequence_number": sequence_number,
                }
                if parent_id is not None:
                    try:
                        pending_ids.add(int(parent_id))
                    except Exception:
                        pass

    _load_family_items_with_ancestors(family_list_ids)

    def _family_root_meta(fid: int):
        cursor = int(fid)
        seen = set()
        root = None
        while cursor and cursor not in seen:
            seen.add(cursor)
            meta = family_item_meta_by_id.get(cursor)
            if not meta:
                break
            root = meta
            parent_id = meta.get("parent_id")
            if not parent_id:
                break
            cursor = int(parent_id)
        return root

    def _family_path(fid: int):
        if not fid:
            return []
        path = []
        cursor = int(fid)
        seen = set()
        while cursor and cursor not in seen:
            seen.add(cursor)
            meta = family_item_meta_by_id.get(cursor)
            if not meta:
                break
            path.append(meta)
            cursor = meta.get("parent_id")
        return list(reversed(path))

    def _parse_standard_type_from_family(level2_meta):
        if not level2_meta:
            return (None, None)
        seq = level2_meta.get("sequence_number")
        name = level2_meta.get("name")
        if seq:
            seq = str(seq).strip() or None
        if name:
            name = str(name).strip() or None
        if seq or name:
            return (seq, name)

        combined = str(level2_meta.get("name") or "").strip()
        if not combined:
            return (None, None)
        import re

        m = re.match(r"^\s*([0-9]+(?:\.[0-9]+)*)\s+(.+?)\s*$", combined)
        if m:
            return (m.group(1), m.group(2))
        m = re.match(r"^\s*([0-9]+)\.(.+?)\s*$", combined)
        if m:
            return (m.group(1), m.group(2).strip())
        return (None, combined)

    def _root_label(root_meta) -> Optional[str]:
        if not root_meta:
            return None
        seq = root_meta.get("sequence_number")
        name = root_meta.get("name")
        if seq and name:
            return f"{str(seq).strip()}.{str(name).strip()}"
        return str(name).strip() if name else None

    def _norm_label(value: Optional[str]) -> str:
        raw = (value or "").strip().lower()
        # keep only a-z/0-9 and '.' so "14. Manual_Input" == "14.manual_input"
        cleaned = "".join(ch for ch in raw if ch.isalnum() or ch == ".")
        return cleaned

    manual_assignment_ids = set()
    for aid, fid in assignment_family_list_id_by_id.items():
        root = _family_root_meta(fid)
        label = _root_label(root)
        if _norm_label(label) == _norm_label("14.Manual_Input"):
            manual_assignment_ids.add(int(aid))

    # Calc dictionary vars per family_list_id
    calc_entries_by_family_list_id = {}
    if family_list_ids:
        calc_entries = (
            db.query(models.CalcDictionaryEntry)
            .filter(models.CalcDictionaryEntry.family_list_id.in_(family_list_ids))
            .order_by(
                models.CalcDictionaryEntry.family_list_id,
                models.CalcDictionaryEntry.symbol_key,
            )
            .all()
        )
        for entry in calc_entries:
            fid = getattr(entry, "family_list_id", None)
            if fid is None:
                continue
            try:
                fid_int = int(fid)
            except Exception:
                continue
            calc_entries_by_family_list_id.setdefault(fid_int, []).append(entry)

    now_iso = datetime.datetime.utcnow().isoformat()

    for cart_entry_id, normalized in cart_entry_payloads:
        aids = normalized.get("assignment_ids") or []
        sids = normalized.get("standard_item_ids") or []
        formula = normalized.get("formula")
        if not formula:
            skipped += 1
            continue

        is_manual = False
        entry_family_list_ids = set()
        for aid in aids:
            try:
                aid_int = int(aid)
            except Exception:
                continue
            if aid_int in manual_assignment_ids:
                is_manual = True
            fid = assignment_family_list_id_by_id.get(aid_int)
            if fid:
                entry_family_list_ids.add(int(fid))
        if not is_manual:
            skipped += 1
            continue

        manual_entries_matched += 1

        # Standard type info: pick a representative family_list_id from the entry
        std_type_number = None
        std_type_name = None
        manual_fid = None
        for aid in aids:
            try:
                aid_int = int(aid)
            except Exception:
                continue
            if aid_int not in manual_assignment_ids:
                continue
            fid = assignment_family_list_id_by_id.get(aid_int)
            if fid:
                manual_fid = int(fid)
                break
        if manual_fid:
            path = _family_path(manual_fid)
            if path:
                level2_meta = path[2] if len(path) > 2 else path[-1]
                std_type_number, std_type_name = _parse_standard_type_from_family(
                    level2_meta
                )

        # Selected work master info: use first standard_item_id
        wm_meta = None
        standard_item_id_value = None
        for sid in sids:
            try:
                sid_int = int(sid)
            except Exception:
                continue
            if standard_item_id_value is None:
                standard_item_id_value = sid_int
            wm_meta = selected_work_master_by_standard_item_id.get(sid_int)
            if wm_meta:
                break
        work_master_id = wm_meta.get("work_master_id") if wm_meta else None
        work_master_code = wm_meta.get("work_master_code") if wm_meta else None
        unit_val = _compose_unit_from_wm_meta(wm_meta)

        standard_item_type_value = (
            _coerce_str(standard_item_type_by_id.get(standard_item_id_value))
            if standard_item_id_value is not None
            else None
        )
        standard_tree_level1_value = _standard_tree_level1_name(standard_item_id_value)
        standard_tree_level2_value = _standard_tree_level2_name(standard_item_id_value)
        if standard_tree_level1_value and standard_tree_level2_value:
            detail_classification_value = (
                f"{standard_tree_level1_value} | {standard_tree_level2_value}"
            )
        else:
            detail_classification_value = (
                standard_tree_level2_value
                or standard_tree_level1_value
                or _coerce_str(standard_item_name_by_id.get(standard_item_id_value))
            )

        variables = {}
        for fid in entry_family_list_ids:
            for ce in calc_entries_by_family_list_id.get(fid, []):
                key = getattr(ce, "symbol_key", None)
                val = getattr(ce, "symbol_value", None)
                if not key:
                    continue
                num = _try_parse_float(val)
                if num is None:
                    continue
                variables[str(key).strip()] = num

        result_val = _safe_eval_numeric_expr(str(formula), variables)
        if result_val is None:
            skipped += 1
            continue

        substituted_formula = str(formula)
        result_log = f"manual_update:{now_iso}"

        for bname in building_names:
            key = f"{_sanitize_filename_part(rev_key)}|{_sanitize_filename_part(bname)}|manual|{cart_entry_id}"
            try:
                db.execute(
                    text(
                        """
                        INSERT OR REPLACE INTO calc_result (
                            key, rev_key, building_name, guid, gui, member_name,
                            category, standard_type_number, standard_type_name,
                            classification, detail_classification, unit,
                            formula, substituted_formula, result, result_log,
                            work_master_id, work_master_code, created_at
                        ) VALUES (
                            :key, :rev_key, :building_name, :guid, :gui, :member_name,
                            :category, :standard_type_number, :standard_type_name,
                            :classification, :detail_classification, :unit,
                            :formula, :substituted_formula, :result, :result_log,
                            :work_master_id, :work_master_code, :created_at
                        )
                        """
                    ),
                    {
                        "key": key,
                        "rev_key": rev_key,
                        "building_name": bname,
                        "guid": "수동항목",
                        "gui": None,
                        "member_name": "Manual_Input",
                        "category": "14.Manual_Input",
                        "standard_type_number": std_type_number,
                        "standard_type_name": std_type_name,
                        "classification": standard_item_type_value or "Manual_Input",
                        "detail_classification": detail_classification_value,
                        "unit": unit_val,
                        "formula": str(formula),
                        "substituted_formula": substituted_formula,
                        "result": float(result_val),
                        "result_log": result_log,
                        "work_master_id": work_master_id,
                        "work_master_code": work_master_code,
                        "created_at": now_iso,
                    },
                )
                inserted += 1
            except Exception:
                skipped += 1
                continue

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Failed manual update")

    return {
        "project_identifier": project_identifier,
        "rev_key": rev_key,
        "buildings": len(building_names),
        "cart_entries_scanned": cart_entries_scanned,
        "manual_entries_matched": manual_entries_matched,
        "inserted": inserted,
        "skipped": skipped,
    }


@router.delete(
    "/project/{project_identifier}/calc-result",
    response_model=schemas.CalcResultDeleteResponse,
    tags=["Project Data"],
)
def delete_calc_results_by_revision(
    project_identifier: str,
    building_name: str,
    rev_key: str,
    db: Session = Depends(get_project_db_session),
):
    building_name = _coerce_str(building_name)
    rev_key = _coerce_str(rev_key)
    if not building_name:
        raise HTTPException(status_code=400, detail="building_name is required")
    if not rev_key:
        raise HTTPException(status_code=400, detail="rev_key is required")

    try:
        res = db.execute(
            text(
                "DELETE FROM calc_result WHERE TRIM(COALESCE(building_name,'')) = TRIM(:building_name) AND rev_key = :rev_key"
            ),
            {"building_name": building_name, "rev_key": rev_key},
        )
        deleted = int(getattr(res, "rowcount", 0) or 0)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Failed to delete calc results")

    return {
        "project_identifier": project_identifier,
        "building_name": building_name,
        "rev_key": rev_key,
        "deleted": deleted,
    }


@router.get(
    "/project/{project_identifier}/export/db-excel",
    tags=["Project Data"],
)
def export_project_db_excel(project_identifier: str):
    """Export a human-reviewable Excel report for the project DB.

    NOTE: This is intentionally *not* a raw table dump. It generates joined/flattened
    sheets so a person can review without jumping across tables.
    """

    try:
        db_path = project_db.resolve_project_db_path(project_identifier)
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    project_db.ensure_extra_tables(db_path)

    pjt_abbr = None
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Excel export dependency missing")

    def _safe_sheet_name(name: str, existing: set) -> str:
        base = (name or "Sheet").strip() or "Sheet"
        cleaned = "".join(ch for ch in base if ch not in "[]:*?/\\")
        cleaned = cleaned[:31] or "Sheet"
        candidate = cleaned
        counter = 1
        while candidate in existing:
            suffix = f"_{counter}"
            candidate = (cleaned[: max(0, 31 - len(suffix))] + suffix)[:31]
            counter += 1
        existing.add(candidate)
        return candidate

    def _excel_value(value):
        if value is None:
            return None
        if isinstance(value, bytes):
            try:
                return value.decode("utf-8", errors="replace")
            except Exception:
                return value.hex()
        if isinstance(value, (dict, list)):
            try:
                return json.dumps(value, ensure_ascii=False)
            except Exception:
                return str(value)
        return value

    def _excel_escape_formula(value):
        if value is None:
            return None
        text_value = str(value)
        if not text_value:
            return text_value
        return text_value if text_value.startswith("'") else f"'{text_value}"

    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    sheet_names = set()
    generated_at = datetime.datetime.now().isoformat(timespec="seconds")

    def _write_sheet_from_df(title: str, df: "pd.DataFrame"):
        ws = wb.create_sheet(_safe_sheet_name(title, sheet_names))
        ws.freeze_panes = "A2"

        if df is None or df.empty:
            ws.append(["(no rows)"])
            return ws

        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([_excel_value(v) for v in row.tolist()])

        # Basic width cap to keep file readable
        for idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(idx)
            max_len = len(str(col_name))
            for cell in ws[col_letter]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
                if max_len > 60:
                    max_len = 60
                    break
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

        return ws

    conn = sqlite3.connect(db_path.as_posix())
    try:

        def _natural_key(text: str):
            import re

            value = (text or "").strip().casefold()
            parts = re.split(r"(\d+)", value)
            key = []
            for part in parts:
                if part.isdigit():
                    try:
                        key.append(int(part))
                    except Exception:
                        key.append(part)
                else:
                    key.append(part)
            return key

        def _parse_sequence_identifier(value):
            import re

            trimmed = str(value).strip() if value is not None else ""
            if not trimmed:
                return None
            match = re.match(r"^(\d+(?:\.\d+)*)([a-zA-Z]*)", trimmed)
            if not match:
                return None
            number_parts = []
            for seg in match.group(1).split("."):
                try:
                    number_parts.append(int(seg))
                except Exception:
                    return None
            suffix = (match.group(2) or "").lower()
            return {"numbers": number_parts, "suffix": suffix}

        def _get_sequence_identifier(node):
            return _parse_sequence_identifier(
                node.get("sequence_number")
            ) or _parse_sequence_identifier(node.get("name"))

        def _compare_sequence_identifiers(a, b):
            max_len = max(len(a["numbers"]), len(b["numbers"]))
            for i in range(max_len):
                va = a["numbers"][i] if i < len(a["numbers"]) else 0
                vb = b["numbers"][i] if i < len(b["numbers"]) else 0
                if va != vb:
                    return -1 if va < vb else 1
            if a["suffix"] != b["suffix"]:
                if not a["suffix"]:
                    return -1
                if not b["suffix"]:
                    return 1
                return -1 if a["suffix"] < b["suffix"] else 1
            return 0

        def _compare_family_nodes(a, b):
            ida = _get_sequence_identifier(a)
            idb = _get_sequence_identifier(b)
            if ida and idb:
                c = _compare_sequence_identifiers(ida, idb)
                if c != 0:
                    return c
            elif ida:
                return -1
            elif idb:
                return 1

            name_a = (a.get("name") or "").strip()
            name_b = (b.get("name") or "").strip()
            if not name_a:
                return 1 if name_b else 0
            if not name_b:
                return -1
            ka = _natural_key(name_a)
            kb = _natural_key(name_b)
            if ka == kb:
                return 0
            return -1 if ka < kb else 1

        def _build_family_tree_rows(
            items,
            assignments_by_family_id=None,
            standard_item_name_by_id=None,
            standard_item_type_by_id=None,
            standard_item_parent_by_id=None,
        ):
            from functools import cmp_to_key
            import math

            def _to_int_or_none(value):
                if value is None:
                    return None
                if isinstance(value, float) and math.isnan(value):
                    return None
                try:
                    text = str(value).strip()
                    if not text or text.lower() == "nan":
                        return None
                    return int(float(text))
                except Exception:
                    return None

            node_by_id = {}
            for item in items:
                node_id = _to_int_or_none(item.get("id"))
                if node_id is None:
                    continue
                parent_id = _to_int_or_none(item.get("parent_id"))
                node_by_id[node_id] = {
                    **item,
                    "id": node_id,
                    "parent_id": parent_id,
                    "children": [],
                }

            roots = []
            for node_id, node in node_by_id.items():
                parent_id = node.get("parent_id")
                if parent_id is not None and parent_id in node_by_id:
                    node_by_id[parent_id]["children"].append(node)
                else:
                    roots.append(node)

            def sort_nodes(nodes):
                nodes.sort(key=cmp_to_key(_compare_family_nodes))
                for n in nodes:
                    if n.get("children"):
                        sort_nodes(n["children"])

            sort_nodes(roots)

            rows = []

            assignments_by_family_id = assignments_by_family_id or {}
            standard_item_name_by_id = standard_item_name_by_id or {}
            standard_item_type_by_id = standard_item_type_by_id or {}
            standard_item_parent_by_id = standard_item_parent_by_id or {}

            def _build_standard_item_subtree_rows(family_id: int, base_level: int):
                assigned = assignments_by_family_id.get(family_id, []) or []
                if not assigned:
                    return

                assigned_by_std_id = {}
                for a in assigned:
                    sid = a.get("standard_item_id")
                    try:
                        sid_int = int(sid) if sid is not None else None
                    except Exception:
                        sid_int = None
                    if sid_int is None:
                        continue
                    # Keep the first assignment if duplicates exist.
                    assigned_by_std_id.setdefault(sid_int, a)

                if not assigned_by_std_id:
                    return

                included_ids = set(assigned_by_std_id.keys())
                stack = list(included_ids)
                while stack:
                    sid = stack.pop()
                    pid = standard_item_parent_by_id.get(sid)
                    if pid is None:
                        continue
                    try:
                        pid_int = int(pid)
                    except Exception:
                        continue
                    if pid_int not in included_ids:
                        included_ids.add(pid_int)
                        stack.append(pid_int)

                children_by_parent = {sid: [] for sid in included_ids}
                roots = []
                for sid in included_ids:
                    pid = standard_item_parent_by_id.get(sid)
                    try:
                        pid_int = int(pid) if pid is not None else None
                    except Exception:
                        pid_int = None
                    if pid_int is not None and pid_int in included_ids:
                        children_by_parent[pid_int].append(sid)
                    else:
                        roots.append(sid)

                def _std_sort_key(sid: int):
                    t = (standard_item_type_by_id.get(sid) or "").strip()
                    # Prefer GWM then SWM, then others
                    type_rank = 2
                    if t == "GWM":
                        type_rank = 0
                    elif t == "SWM":
                        type_rank = 1
                    name = (standard_item_name_by_id.get(sid) or "").strip()
                    return (type_rank, _natural_key(name), sid)

                for pid in list(children_by_parent.keys()):
                    children_by_parent[pid].sort(key=_std_sort_key)
                roots.sort(key=_std_sort_key)

                def walk_std(sid: int, depth: int):
                    name = standard_item_name_by_id.get(sid)
                    typ = standard_item_type_by_id.get(sid)
                    assignment = assigned_by_std_id.get(sid)

                    pid = standard_item_parent_by_id.get(sid)
                    parent_id = None
                    if pid is not None:
                        try:
                            pid_int = int(pid)
                        except Exception:
                            pid_int = None
                        if pid_int is not None and pid_int in included_ids:
                            parent_id = pid_int
                    if parent_id is None:
                        parent_id = family_id

                    rows.append(
                        {
                            "level": int(base_level) + 1 + int(depth),
                            "sequence_number": None,
                            "name": name,
                            "item_type": typ,
                            "id": sid,
                            "parent_id": parent_id,
                            "description": (
                                assignment.get("assignment_description")
                                if assignment
                                else None
                            ),
                            "formula": (
                                _excel_escape_formula(assignment.get("formula"))
                                if assignment
                                else None
                            ),
                            "created_at": (
                                (
                                    assignment.get("assigned_at")
                                    or assignment.get("created_at")
                                )
                                if assignment
                                else None
                            ),
                        }
                    )
                    for child_id in children_by_parent.get(sid, []) or []:
                        walk_std(child_id, depth + 1)

                for root_id in roots:
                    walk_std(root_id, 0)

            def walk(nodes, level=0):
                for n in nodes:
                    seq = (n.get("sequence_number") or "").strip()
                    name = (n.get("name") or "").strip() or "Unnamed"
                    family_id = n.get("id")
                    rows.append(
                        {
                            "level": int(level),
                            "sequence_number": seq or None,
                            "name": name,
                            "item_type": n.get("item_type"),
                            "id": n.get("id"),
                            "parent_id": n.get("parent_id"),
                            "description": n.get("description"),
                            "formula": None,
                            "created_at": n.get("created_at"),
                        }
                    )

                    # Insert assigned standard items right under this family node, preserving hierarchy.
                    try:
                        family_id_int = (
                            int(family_id) if family_id is not None else None
                        )
                    except Exception:
                        family_id_int = None
                    if family_id_int is not None:
                        _build_standard_item_subtree_rows(
                            family_id=family_id_int,
                            base_level=int(level),
                        )

                    children = n.get("children") or []
                    if children:
                        walk(children, level + 1)

            walk(roots, 0)
            return rows

        def _read_df(query: str, params=None) -> "pd.DataFrame":
            try:
                return pd.read_sql_query(query, conn, params=params)
            except Exception:
                return pd.DataFrame()

        # Summary / metadata
        summary_ws = wb.create_sheet(_safe_sheet_name("SUMMARY", sheet_names))
        summary_ws.append(["project_identifier", project_identifier])
        summary_ws.append(["db_file", db_path.name])
        summary_ws.append(["generated_at", generated_at])
        summary_ws.append([])
        summary_ws.append(["sheet", "rows"])

        # Report_WM (WM pre-check)
        df_wm_precheck_raw = _read_df(
            """
            SELECT
              wm.id AS work_master_id,
              COALESCE(wmp.use_yn, 0) AS use_yn,
                            wmp.other_opinion,
              wm.work_master_code,
              wm.gauge,
              wm.uom1,
              wm.uom2,
              wm.add_spec,
              wm.discipline,
              wm.cat_large_code,
              wm.cat_large_desc,
              wm.cat_mid_code,
              wm.cat_mid_desc,
              wm.cat_small_code,
              wm.cat_small_desc,
              wm.attr1_code,
              wm.attr1_spec,
              wm.attr2_code,
              wm.attr2_spec,
              wm.attr3_code,
              wm.attr3_spec,
              wm.attr4_code,
              wm.attr4_spec,
              wm.attr5_code,
              wm.attr5_spec,
              wm.attr6_code,
              wm.attr6_spec,
              wm.work_group_code,
              wm.new_old_code,
              wmp.updated_at
            FROM work_masters wm
            LEFT JOIN work_master_precheck wmp ON wmp.work_master_id = wm.id
            WHERE LOWER(COALESCE(TRIM(wm.new_old_code), '')) <> 'old'
              AND (wm.work_master_code NOT LIKE 'S%' OR COALESCE(wm.cat_mid_code, '') = 'AA')
              AND (wm.work_master_code NOT LIKE 'F%' OR COALESCE(wm.cat_large_code, '') = 'F01')
            ORDER BY wm.work_master_code, wm.gauge, wm.id
            """
        )

        def _wm_trim(value):
            return str(value).strip() if value is not None else ""

        def _wm_summary_parts(row) -> list:
            parts = []

            def add(label, value):
                v = _wm_trim(value)
                if not v:
                    return
                parts.append((label, v))

            add("Discipline", row.get("discipline"))
            add(
                "Large",
                " ".join(
                    [
                        _wm_trim(row.get("cat_large_code")),
                        _wm_trim(row.get("cat_large_desc")),
                    ]
                ).strip(),
            )
            add(
                "Mid",
                " ".join(
                    [
                        _wm_trim(row.get("cat_mid_code")),
                        _wm_trim(row.get("cat_mid_desc")),
                    ]
                ).strip(),
            )
            add(
                "Small",
                " ".join(
                    [
                        _wm_trim(row.get("cat_small_code")),
                        _wm_trim(row.get("cat_small_desc")),
                    ]
                ).strip(),
            )
            add(
                "Attr1",
                " ".join(
                    [_wm_trim(row.get("attr1_code")), _wm_trim(row.get("attr1_spec"))]
                ).strip(),
            )
            add(
                "Attr2",
                " ".join(
                    [_wm_trim(row.get("attr2_code")), _wm_trim(row.get("attr2_spec"))]
                ).strip(),
            )
            add(
                "Attr3",
                " ".join(
                    [_wm_trim(row.get("attr3_code")), _wm_trim(row.get("attr3_spec"))]
                ).strip(),
            )
            add(
                "Attr4",
                " ".join(
                    [_wm_trim(row.get("attr4_code")), _wm_trim(row.get("attr4_spec"))]
                ).strip(),
            )
            add(
                "Attr5",
                " ".join(
                    [_wm_trim(row.get("attr5_code")), _wm_trim(row.get("attr5_spec"))]
                ).strip(),
            )
            add(
                "Attr6",
                " ".join(
                    [_wm_trim(row.get("attr6_code")), _wm_trim(row.get("attr6_spec"))]
                ).strip(),
            )
            add("Group", row.get("work_group_code"))
            add("New/Old", row.get("new_old_code"))

            return parts

        df_wm_precheck = (
            df_wm_precheck_raw.copy()
            if df_wm_precheck_raw is not None
            else pd.DataFrame()
        )

        selected_work_master_ids = set()
        try:
            df_selected = _read_df(
                "SELECT work_master_id FROM standard_item_work_master_select"
            )
            if df_selected is not None and not df_selected.empty:
                for v in df_selected["work_master_id"].tolist():
                    try:
                        selected_work_master_ids.add(int(v))
                    except Exception:
                        continue
        except Exception:
            selected_work_master_ids = set()

        selected_row_flags = []
        if df_wm_precheck is not None and not df_wm_precheck.empty:
            ui_use = []
            ui_code = []
            ui_gauge = []
            ui_unit = []
            ui_spec = []
            ui_other_opinion = []
            ui_work_master = []
            for _, r in df_wm_precheck.iterrows():
                try:
                    wm_id = r.get("work_master_id")
                    selected_row_flags.append(
                        int(wm_id) in selected_work_master_ids
                        if wm_id is not None
                        else False
                    )
                except Exception:
                    selected_row_flags.append(False)

                wm_code = _wm_trim(r.get("work_master_code"))
                gauge_value = _wm_trim(r.get("gauge")).upper()
                wm_title = (
                    (f"{wm_code}({gauge_value})" if gauge_value else wm_code)
                    if wm_code
                    else (f"({gauge_value})" if gauge_value else "코드 정보 없음")
                )

                headline = (
                    _wm_trim(r.get("cat_large_desc"))
                    or _wm_trim(r.get("cat_mid_desc"))
                    or _wm_trim(r.get("cat_small_desc"))
                    or wm_title
                )

                unit_label = " / ".join(
                    [v for v in [_wm_trim(r.get("uom1")), _wm_trim(r.get("uom2"))] if v]
                )
                spec_value = str(r.get("add_spec") or "")
                other_opinion_value = str(r.get("other_opinion") or "")

                parts = _wm_summary_parts(r)
                summary = " | ".join([f"{k}={v}" for k, v in parts])
                work_master_cell = f"{headline}\n{wm_title}" + (
                    f"\n{summary}" if summary else ""
                )

                ui_use.append(bool(r.get("use_yn")))
                ui_code.append(wm_code)
                ui_gauge.append(gauge_value)
                ui_unit.append(unit_label)
                ui_spec.append(spec_value)
                ui_other_opinion.append(other_opinion_value)
                ui_work_master.append(work_master_cell)

            # Insert UI columns first (same order as WM pre-check table)
            df_wm_precheck.insert(0, "Work Master", ui_work_master)
            df_wm_precheck.insert(0, "기타의견", ui_other_opinion)
            df_wm_precheck.insert(0, "Spec", ui_spec)
            df_wm_precheck.insert(0, "Unit", ui_unit)
            df_wm_precheck.insert(0, "Gauge", ui_gauge)
            df_wm_precheck.insert(0, "WM Code", ui_code)
            df_wm_precheck.insert(0, "Use", ui_use)

            # Omit columns from `work_master_id` and everything to the right.
            if "work_master_id" in df_wm_precheck.columns:
                keep_end = int(df_wm_precheck.columns.get_loc("work_master_id"))
                df_wm_precheck = df_wm_precheck.loc[
                    :, df_wm_precheck.columns[:keep_end]
                ]

        ws_wm = _write_sheet_from_df("Report_WM", df_wm_precheck)

        # Improve readability for the UI-style text columns.
        try:
            header_cells = list(
                ws_wm.iter_rows(min_row=1, max_row=1, values_only=False)
            )[0]
            header_to_col = {c.value: c.column for c in header_cells}
            for header_name in ("Spec", "기타의견", "Work Master"):
                col_idx = header_to_col.get(header_name)
                if not col_idx:
                    continue
                col_letter = get_column_letter(col_idx)
                ws_wm.column_dimensions[col_letter].width = (
                    60 if header_name == "Work Master" else 40
                )
                for cell in ws_wm[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

            # --- WM pre-check screen color cues (approximation) ---
            # - Gauge text is purple & bold
            # - WM Code cell is bold, and highlighted when selected in Standard Select

            header_fill = PatternFill("solid", fgColor="FFF9FAFB")
            header_font = Font(bold=True)
            gauge_font = Font(color="FF9333EA", bold=True)
            wm_code_font = Font(bold=True)
            wm_selected_fill = PatternFill("solid", fgColor="FFEDE9FE")
            wm_selected_font = Font(color="FF4C1D95", bold=True)

            # Header styling
            for cell in ws_wm[1]:
                cell.fill = header_fill
                cell.font = header_font

            col_use = header_to_col.get("Use")
            col_wm_code = header_to_col.get("WM Code")
            col_gauge = header_to_col.get("Gauge")

            if col_use:
                col_letter = get_column_letter(col_use)
                for cell in ws_wm[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Row-wise styling (keep it minimal for performance)
            if col_wm_code or col_gauge:
                max_row = ws_wm.max_row
                for r in range(2, max_row + 1):
                    if col_gauge:
                        c = ws_wm.cell(row=r, column=col_gauge)
                        c.font = gauge_font

                    if col_wm_code:
                        c = ws_wm.cell(row=r, column=col_wm_code)
                        selected = False
                        try:
                            idx = r - 2
                            if idx >= 0 and idx < len(selected_row_flags):
                                selected = bool(selected_row_flags[idx])
                        except Exception:
                            selected = False

                        if selected:
                            c.fill = wm_selected_fill
                            c.font = wm_selected_font
                        else:
                            c.font = wm_code_font

            # Ensure all cells are vertically centered (including header).
            for row in ws_wm.iter_rows():
                for cell in row:
                    try:
                        cell.alignment = cell.alignment.copy(vertical="center")
                    except Exception:
                        cell.alignment = Alignment(vertical="center")

            # Scale down font size (~80%).
            scale = 0.8
            default_font_size = 11

            for row in ws_wm.iter_rows():
                for cell in row:
                    try:
                        base = (
                            cell.font.size
                            if cell.font.size is not None
                            else default_font_size
                        )
                        cell.font = cell.font.copy(size=max(1, base * scale))
                    except Exception:
                        cell.font = Font(size=max(1, default_font_size * scale))

            default_row_height = 15
            base_line_height = default_row_height * scale

            def _line_count(value) -> int:
                if value is None:
                    return 1
                try:
                    text = str(value)
                except Exception:
                    return 1
                return max(1, text.count("\n") + 1)

            for r in range(1, ws_wm.max_row + 1):
                try:
                    max_lines = 1
                    for c in range(1, ws_wm.max_column + 1):
                        v = ws_wm.cell(row=r, column=c).value
                        max_lines = max(max_lines, _line_count(v))
                    ws_wm.row_dimensions[r].height = max(
                        1, base_line_height * max_lines
                    )
                except Exception:
                    continue
        except Exception:
            pass

        summary_ws.append(["Report_WM", int(len(df_wm_precheck.index))])

        # Family list (tree, as shown in app) + assigned standard items under each node.
        df_family_raw = _read_df(
            "SELECT id, parent_id, sequence_number, name, item_type, description, created_at FROM family_list ORDER BY id"
        )
        df_family_assignments = _read_df(
            """
            SELECT
              g.id AS assignment_id,
              g.family_list_id,
              g.standard_item_id,
              si.name AS standard_item_name,
              si.type AS standard_item_type,
              g.formula,
              g.description AS assignment_description,
              g.assigned_at,
              g.created_at
            FROM gwm_family_assign g
            LEFT JOIN standard_items si ON si.id = g.standard_item_id
            ORDER BY g.family_list_id, si.type, si.name, g.id
            """
        )

        # Project abbreviation (for derived item name formatting)
        pjt_abbr = None
        try:
            df_meta = _read_df(
                "SELECT pjt_abbr FROM project_metadata ORDER BY id LIMIT 1"
            )
            if df_meta is not None and not df_meta.empty:
                raw_abbr = df_meta.iloc[0].get("pjt_abbr")
                if raw_abbr is not None:
                    abbr = str(raw_abbr).strip()
                    pjt_abbr = abbr or None
        except Exception:
            pjt_abbr = None

        # Standard item hierarchy (for indentation of assigned items)
        df_standard_items_hier = _read_df(
            "SELECT id, parent_id, derive_from, name AS standard_item_name, type AS standard_item_type FROM standard_items"
        )
        standard_item_name_by_id = {}
        standard_item_type_by_id = {}
        standard_item_parent_by_id = {}
        standard_item_derive_from_by_id = {}
        if df_standard_items_hier is not None and not df_standard_items_hier.empty:
            for _, r in df_standard_items_hier.iterrows():
                sid = r.get("id")
                try:
                    sid_int = int(sid) if sid is not None else None
                except Exception:
                    sid_int = None
                if sid_int is None:
                    continue
                standard_item_name_by_id[sid_int] = r.get("standard_item_name")
                standard_item_type_by_id[sid_int] = r.get("standard_item_type")
                pid = r.get("parent_id")
                try:
                    pid_int = int(pid) if pid is not None else None
                except Exception:
                    pid_int = None
                standard_item_parent_by_id[sid_int] = pid_int

                derive_from = r.get("derive_from")
                try:
                    derive_from_int = (
                        int(derive_from) if derive_from is not None else None
                    )
                except Exception:
                    derive_from_int = None
                standard_item_derive_from_by_id[sid_int] = derive_from_int

        # Apply derived item naming: sourceName [abbr]::baseName
        if standard_item_derive_from_by_id:
            for sid_int, source_id in list(standard_item_derive_from_by_id.items()):
                if source_id is None:
                    continue
                base_name = standard_item_name_by_id.get(sid_int)
                source_name = standard_item_name_by_id.get(source_id)
                if not base_name or not source_name:
                    continue
                base_name = str(base_name).replace("\u00a0", " ").strip()
                source_name = str(source_name).replace("\u00a0", " ").strip()
                if not base_name or not source_name:
                    continue
                if pjt_abbr:
                    formatted = f"{source_name} [{pjt_abbr}]::{base_name}"
                else:
                    formatted = f"{source_name}::{base_name}"

                # Enforce no whitespace after '::'
                while ":: " in formatted:
                    formatted = formatted.replace(":: ", "::")
                formatted = formatted.replace("::\u00a0", "::")
                standard_item_name_by_id[sid_int] = formatted
        assignments_by_family_id = {}
        if df_family_assignments is not None and not df_family_assignments.empty:
            for _, r in df_family_assignments.iterrows():
                fid = r.get("family_list_id")
                try:
                    fid_int = int(fid) if fid is not None else None
                except Exception:
                    fid_int = None
                if fid_int is None:
                    continue
                assignments_by_family_id.setdefault(fid_int, []).append(
                    {
                        "assignment_id": r.get("assignment_id"),
                        "standard_item_id": r.get("standard_item_id"),
                        "standard_item_name": r.get("standard_item_name"),
                        "standard_item_type": r.get("standard_item_type"),
                        "formula": r.get("formula"),
                        "assignment_description": r.get("assignment_description"),
                        "assigned_at": r.get("assigned_at"),
                        "created_at": r.get("created_at"),
                    }
                )

        family_rows = _build_family_tree_rows(
            df_family_raw.to_dict(orient="records") if not df_family_raw.empty else [],
            assignments_by_family_id=assignments_by_family_id,
            standard_item_name_by_id=standard_item_name_by_id,
            standard_item_type_by_id=standard_item_type_by_id,
            standard_item_parent_by_id=standard_item_parent_by_id,
        )
        df_family_tree = pd.DataFrame(family_rows)
        if df_family_tree is not None and not df_family_tree.empty:
            df_family_tree = df_family_tree.drop(
                columns=["created_at"], errors="ignore"
            )
            preferred_cols = [
                "level",
                "sequence_number",
                "name",
                "item_type",
                "id",
                "parent_id",
                "formula",
                "description",
            ]
            existing = [c for c in preferred_cols if c in df_family_tree.columns]
            remainder = [
                c
                for c in df_family_tree.columns
                if c not in set(existing) and c != "created_at"
            ]
            df_family_tree = df_family_tree[existing + remainder]

        family_ws = _write_sheet_from_df("Report_FamilyList", df_family_tree)
        if family_ws and df_family_tree is not None and not df_family_tree.empty:
            headers = [cell.value for cell in family_ws[1]]
            try:
                level_col = headers.index("level") + 1
                name_col = headers.index("name") + 1
            except ValueError:
                level_col = None
                name_col = None

            if level_col and name_col:
                for row_idx in range(2, family_ws.max_row + 1):
                    level_value = family_ws.cell(row=row_idx, column=level_col).value
                    try:
                        indent_level = int(level_value or 0)
                    except Exception:
                        indent_level = 0
                    name_cell = family_ws.cell(row=row_idx, column=name_col)
                    name_cell.alignment = Alignment(indent=indent_level, wrap_text=True)
        summary_ws.append(["Report_FamilyList", int(len(df_family_tree.index))])

        df_buildings = _read_df(
            "SELECT id, name AS building_name, created_at FROM building_list ORDER BY id"
        )
        _write_sheet_from_df("Buildings", df_buildings)
        summary_ws.append(["Buildings", int(len(df_buildings.index))])

        df_standard_items = _read_df(
            "SELECT id, name AS standard_item_name, type AS standard_item_type, parent_id, derive_from FROM standard_items ORDER BY id"
        )
        _write_sheet_from_df("StandardItems", df_standard_items)
        summary_ws.append(["StandardItems", int(len(df_standard_items.index))])

        df_work_masters = _read_df(
            "SELECT id, discipline, work_master_code, cat_large_code, cat_large_desc, cat_mid_code, cat_mid_desc, cat_small_code, cat_small_desc, attr1_code, attr1_spec, attr2_code, attr2_spec, attr3_code, attr3_spec, attr4_code, attr4_spec, attr5_code, attr5_spec, attr6_code, attr6_spec, uom1, uom2, work_group_code, new_old_code, add_spec, gauge FROM work_masters ORDER BY id"
        )
        _write_sheet_from_df("WorkMasters", df_work_masters)
        summary_ws.append(["WorkMasters", int(len(df_work_masters.index))])

        df_selected = _read_df(
            """
            SELECT
              sel.id,
              sel.standard_item_id,
              si.name AS standard_item_name,
              si.type AS standard_item_type,
              sel.work_master_id,
              wm.work_master_code,
              wm.cat_large_code,
              wm.cat_mid_code,
              wm.cat_small_code,
              sel.created_at,
              sel.updated_at
            FROM standard_item_work_master_select sel
            LEFT JOIN standard_items si ON si.id = sel.standard_item_id
            LEFT JOIN work_masters wm ON wm.id = sel.work_master_id
            ORDER BY sel.standard_item_id
            """
        )
        _write_sheet_from_df("StandardItemSelections", df_selected)
        summary_ws.append(["StandardItemSelections", int(len(df_selected.index))])

        df_gwm_assign = _read_df(
            """
            SELECT
              g.id,
              g.family_list_id,
              fl.name AS family_name,
              g.standard_item_id,
              si.name AS standard_item_name,
              si.type AS standard_item_type,
              g.formula,
              g.description,
              g.assigned_at,
              g.created_at
            FROM gwm_family_assign g
            LEFT JOIN family_list fl ON fl.id = g.family_list_id
            LEFT JOIN standard_items si ON si.id = g.standard_item_id
            ORDER BY g.id
            """
        )
        if not df_gwm_assign.empty and "formula" in df_gwm_assign.columns:
            df_gwm_assign["formula"] = df_gwm_assign["formula"].map(
                _excel_escape_formula
            )
        _write_sheet_from_df("GwmFamilyAssign", df_gwm_assign)
        summary_ws.append(["GwmFamilyAssign", int(len(df_gwm_assign.index))])

        df_revit_types = _read_df(
            """
            SELECT
              frt.id,
              frt.family_list_id,
              fl.name AS family_name,
              frt.type_name,
              frt.building_name,
              frt.created_at
            FROM family_revit_type frt
            LEFT JOIN family_list fl ON fl.id = frt.family_list_id
            ORDER BY frt.id
            """
        )
        _write_sheet_from_df("FamilyRevitTypes", df_revit_types)
        summary_ws.append(["FamilyRevitTypes", int(len(df_revit_types.index))])

        df_calc = _read_df(
            """
            SELECT
              c.id,
              c.family_list_id,
              fl.name AS family_name,
              c.calc_code,
              c.symbol_key,
              c.symbol_value,
              c.created_at
            FROM calc_dictionary c
            LEFT JOIN family_list fl ON fl.id = c.family_list_id
            WHERE COALESCE(c.is_deleted, 0) = 0
            ORDER BY c.id
            """
        )
        _write_sheet_from_df("CalcDictionary", df_calc)
        summary_ws.append(["CalcDictionary", int(len(df_calc.index))])

        # Cart entries (flattened for review)
        df_cart_raw = _read_df(
            "SELECT id AS cart_entry_id, payload, created_at FROM workmaster_cart_entries ORDER BY id DESC"
        )
        std_name_by_id = (
            df_standard_items.set_index("id")["standard_item_name"].to_dict()
            if not df_standard_items.empty and "id" in df_standard_items.columns
            else {}
        )
        std_type_by_id = (
            df_standard_items.set_index("id")["standard_item_type"].to_dict()
            if not df_standard_items.empty and "id" in df_standard_items.columns
            else {}
        )
        sel_by_std_id = {}
        if not df_selected.empty and "standard_item_id" in df_selected.columns:
            for _, r in df_selected.iterrows():
                sid = r.get("standard_item_id")
                if sid is None:
                    continue
                sel_by_std_id[int(sid)] = {
                    "selected_work_master_id": r.get("work_master_id"),
                    "selected_work_master_code": r.get("work_master_code"),
                }

        cart_rows = []
        for _, r in df_cart_raw.iterrows():
            raw_payload = r.get("payload")
            try:
                payload = json.loads(raw_payload or "{}")
            except Exception:
                payload = {}
            normalized = _normalize_cart_payload(
                payload if isinstance(payload, dict) else {}
            )
            revit_types = normalized.get("revit_types") or []
            assignment_ids = normalized.get("assignment_ids") or []
            standard_item_ids = normalized.get("standard_item_ids") or []
            building_names = normalized.get("building_names") or []

            standard_item_id = None
            try:
                standard_item_id = (
                    int(standard_item_ids[0]) if standard_item_ids else None
                )
            except Exception:
                standard_item_id = None

            sel = (
                sel_by_std_id.get(standard_item_id)
                if standard_item_id is not None
                else None
            )
            cart_rows.append(
                {
                    "cart_entry_id": r.get("cart_entry_id"),
                    "created_at": r.get("created_at"),
                    "building_name": (building_names[0] if building_names else None),
                    "standard_item_id": standard_item_id,
                    "standard_item_name": std_name_by_id.get(standard_item_id),
                    "standard_item_type": std_type_by_id.get(standard_item_id),
                    "assignment_id": (assignment_ids[0] if assignment_ids else None),
                    "revit_type": (revit_types[0] if revit_types else None),
                    "formula": _excel_escape_formula(normalized.get("formula")),
                    "selected_work_master_id": (
                        sel.get("selected_work_master_id") if sel else None
                    ),
                    "selected_work_master_code": (
                        sel.get("selected_work_master_code") if sel else None
                    ),
                    "building_names_json": json.dumps(
                        building_names, ensure_ascii=False
                    ),
                    "standard_item_ids_json": json.dumps(
                        standard_item_ids, ensure_ascii=False
                    ),
                    "assignment_ids_json": json.dumps(
                        assignment_ids, ensure_ascii=False
                    ),
                    "revit_types_json": json.dumps(revit_types, ensure_ascii=False),
                }
            )
        df_cart = pd.DataFrame(cart_rows)
        _write_sheet_from_df("CartEntries", df_cart)
        summary_ws.append(["CartEntries", int(len(df_cart.index))])
        try:
            cur = conn.cursor()
            cur.execute("SELECT pjt_abbr FROM project_metadata ORDER BY id LIMIT 1")
            row = cur.fetchone()
            if row and row[0]:
                pjt_abbr = str(row[0]).strip() or None
        except Exception:
            pjt_abbr = None
    finally:
        conn.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.datetime.now()
    stamp = now.strftime("%Y%m%d_%H%M%S")
    safe_abbr = (pjt_abbr or project_identifier or "project").strip()
    safe_abbr = (
        safe_abbr.replace("\\", "_")
        .replace("/", "_")
        .replace(":", "_")
        .replace("*", "_")
        .replace("?", "_")
        .replace('"', "_")
        .replace("<", "_")
        .replace(">", "_")
        .replace("|", "_")
    )
    filename = f"DB_{safe_abbr}_{stamp}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post(
    "/project/{project_identifier}/standard-items/{standard_item_id}/derive",
    response_model=schemas.StandardItem,
    tags=["Project Data"],
)
def derive_project_standard_item_work_master(
    project_identifier: str,
    standard_item_id: int,
    payload: schemas.DerivedStandardItemCreate,
    db: Session = Depends(get_project_db_session),
):
    derived = crud.create_derived_standard_item(
        db,
        parent_id=standard_item_id,
        suffix_description=payload.suffix_description,
        work_master_id=payload.work_master_id,
    )
    if not derived:
        raise HTTPException(status_code=400, detail="Derived item could not be created")
    return derived


@router.delete(
    "/project/{project_identifier}/standard-items/{standard_item_id}",
    tags=["Project Data"],
)
def delete_project_standard_item(
    project_identifier: str,
    standard_item_id: int,
    db: Session = Depends(get_project_db_session),
):
    std = crud.delete_standard_item(db, standard_item_id=standard_item_id)
    if not std:
        raise HTTPException(status_code=404, detail="StandardItem not found")
    return {"message": "deleted", "standard_item_id": standard_item_id}


@router.post(
    "/project/{project_identifier}/standard-items/{standard_item_id}/rename",
    tags=["Project Data"],
)
def rename_project_standard_item(
    project_identifier: str,
    standard_item_id: int,
    payload: schemas.StandardItemRename,
    db: Session = Depends(get_project_db_session),
):
    std = crud.update_standard_item_name(
        db, standard_item_id=standard_item_id, new_name=payload.name
    )
    if not std:
        raise HTTPException(status_code=404, detail="StandardItem not found")
    return {"message": "renamed", "standard_item_id": std.id}


@router.post(
    "/project/{project_identifier}/standard-items/",
    response_model=schemas.StandardItem,
    tags=["Project Data"],
)
def create_project_standard_item(
    project_identifier: str,
    item: schemas.StandardItemCreate,
    db: Session = Depends(get_project_db_session),
):
    return crud.create_standard_item(db=db, standard_item=item)


@router.get(
    "/project/{project_identifier}/building-list/",
    response_model=List[schemas.BuildingItem],
    tags=["Project Data"],
)
def read_project_buildings(
    project_identifier: str,
    db: Session = Depends(get_project_db_session),
):
    return crud.list_buildings(db)


@router.post(
    "/project/{project_identifier}/building-list/",
    response_model=schemas.BuildingItem,
    tags=["Project Data"],
)
def create_project_building(
    project_identifier: str,
    building: schemas.BuildingCreate,
    db: Session = Depends(get_project_db_session),
):
    existing = (
        db.query(models.BuildingList)
        .filter(models.BuildingList.name == building.name)
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="Building already exists")
    return crud.create_building(db=db, building=building)


@router.get(
    "/project/{project_identifier}/metadata/abbr",
    response_model=schemas.ProjectMetadata,
    tags=["Project Data"],
)
def read_project_metadata(
    project_identifier: str,
):
    try:
        db_path = project_db.resolve_project_db_path(project_identifier)
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return project_db.read_project_metadata(db_path)


@router.patch(
    "/project/{project_identifier}/metadata/abbr",
    response_model=schemas.ProjectMetadata,
    tags=["Project Data"],
)
def update_project_metadata(
    project_identifier: str,
    payload: schemas.ProjectMetadata,
):
    try:
        db_path = project_db.resolve_project_db_path(project_identifier)
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    updates = payload.model_dump(exclude_unset=True)
    return project_db.update_project_metadata(db_path, updates)


@router.delete(
    "/project/{project_identifier}/building-list/{building_id}",
    tags=["Project Data"],
)
def delete_project_building(
    project_identifier: str,
    building_id: int,
    db: Session = Depends(get_project_db_session),
):
    deleted = crud.delete_building(db=db, building_id=building_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Building not found")
    return {"message": "deleted", "building_id": building_id}


# Get single standard item
@router.get(
    "/standard-items/{standard_item_id}",
    response_model=schemas.StandardItem,
    tags=["Standard Items"],
)
def get_standard_item(standard_item_id: int, db: Session = Depends(get_db)):
    std = crud.get_standard_item(db, standard_item_id=standard_item_id)
    if not std:
        raise HTTPException(status_code=404, detail="StandardItem not found")
    return std


# Delete standard item
@router.delete("/standard-items/{standard_item_id}", tags=["Standard Items"])
def delete_standard_item(standard_item_id: int, db: Session = Depends(get_db)):
    std = crud.delete_standard_item(db, standard_item_id=standard_item_id)
    if not std:
        raise HTTPException(status_code=404, detail="StandardItem not found")
    return {"message": "deleted", "standard_item_id": standard_item_id}


# Rename standard item (only name)
@router.post("/standard-items/{standard_item_id}/rename", tags=["Standard Items"])
def rename_standard_item(
    standard_item_id: int,
    payload: schemas.StandardItemRename,
    db: Session = Depends(get_db),
):
    std = crud.update_standard_item_name(
        db, standard_item_id=standard_item_id, new_name=payload.name
    )
    if not std:
        raise HTTPException(status_code=404, detail="StandardItem not found")
    return {"message": "renamed", "standard_item_id": std.id}


@router.get(
    "/family-list/",
    response_model=List[schemas.FamilyListItem],
    tags=["Family List"],
)
def read_family_list(db: Session = Depends(get_db)):
    return crud.list_family_items(db)


@router.post(
    "/family-list/",
    response_model=schemas.FamilyListItem,
    tags=["Family List"],
)
def create_family_list_item(
    item: schemas.FamilyListCreate, db: Session = Depends(get_db)
):
    return crud.create_family_item(db, item)


@router.put(
    "/family-list/{item_id}",
    response_model=schemas.FamilyListItem,
    tags=["Family List"],
)
def update_family_list_item(
    item_id: int, payload: schemas.FamilyListUpdate, db: Session = Depends(get_db)
):
    updated = crud.update_family_item(db, item_id, payload)
    if not updated:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return updated


@router.delete("/family-list/{item_id}", tags=["Family List"])
def delete_family_list_item(item_id: int, db: Session = Depends(get_db)):
    deleted = crud.delete_family_item(db, item_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return {"message": "deleted", "id": item_id}


@router.get(
    "/family-list/{item_id}/calc-dictionary",
    response_model=List[schemas.CalcDictionaryEntry],
    tags=["Family List"],
)
def read_family_calc_dictionary(item_id: int, db: Session = Depends(get_db)):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.list_calc_dictionary_entries(db, family_item_id=item_id)


@router.post(
    "/family-list/{item_id}/calc-dictionary",
    response_model=schemas.CalcDictionaryEntry,
    tags=["Family List"],
)
def create_family_calc_dictionary_entry(
    item_id: int,
    entry: schemas.CalcDictionaryEntryCreate,
    db: Session = Depends(get_db),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.create_calc_dictionary_entry(db, family_item_id=item_id, entry_in=entry)


@router.patch(
    "/family-list/{item_id}/calc-dictionary/{entry_id}",
    response_model=schemas.CalcDictionaryEntry,
    tags=["Family List"],
)
def update_family_calc_dictionary_entry(
    item_id: int,
    entry_id: int,
    payload: schemas.CalcDictionaryEntryUpdate,
    db: Session = Depends(get_db),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    entry = crud.get_calc_dictionary_entry(db, entry_id)
    if not entry or entry.family_list_id != item_id:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    updates = payload.model_dump(exclude_none=True)
    if not updates:
        return entry
    updated = crud.update_calc_dictionary_entry(db, entry_id=entry_id, updates=updates)
    if not updated:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    return updated


@router.delete(
    "/family-list/{item_id}/calc-dictionary/{entry_id}",
    tags=["Family List"],
)
def delete_family_calc_dictionary_entry(
    item_id: int,
    entry_id: int,
    db: Session = Depends(get_db),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    entry = crud.get_calc_dictionary_entry(db, entry_id)
    if not entry or entry.family_list_id != item_id:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    deleted = crud.delete_calc_dictionary_entry(db, entry_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    return {"message": "deleted", "id": entry_id}


@router.get(
    "/calc-dictionary",
    response_model=List[schemas.CalcDictionaryEntry],
    tags=["Calc Dictionary"],
)
def read_calc_dictionary_index(db: Session = Depends(get_db)):
    return crud.list_all_calc_dictionary_entries(db)


@router.patch(
    "/calc-dictionary/{entry_id}",
    response_model=schemas.CalcDictionaryEntry,
    tags=["Calc Dictionary"],
)
def update_calc_dictionary_entry(
    entry_id: int,
    payload: schemas.CalcDictionaryEntryUpdate,
    db: Session = Depends(get_db),
):
    entry = crud.get_calc_dictionary_entry(db, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    updates = payload.model_dump(exclude_none=True)
    if not updates:
        return entry
    updated = crud.update_calc_dictionary_entry(db, entry_id=entry_id, updates=updates)
    if not updated:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    return updated


@router.post(
    "/calc-dictionary/sync-with-common-input",
    response_model=schemas.CalcDictionarySyncResult,
    tags=["Calc Dictionary"],
)
def sync_calc_dictionary_with_common_input(db: Session = Depends(get_db)):
    updated_entries = crud.sync_calc_dictionary_with_common_inputs(db)
    return schemas.CalcDictionarySyncResult(updated_entries=updated_entries)


@router.get(
    "/project/{project_identifier}/common-input/",
    response_model=List[schemas.CommonInputItem],
    tags=["Project Data"],
)
def list_project_common_input(
    project_identifier: str, db: Session = Depends(get_project_db_session)
):
    return crud.list_common_inputs(db)


@router.post(
    "/project/{project_identifier}/common-input/",
    response_model=schemas.CommonInputItem,
    tags=["Project Data"],
)
def create_project_common_input(
    project_identifier: str,
    payload: schemas.CommonInputCreate,
    db: Session = Depends(get_project_db_session),
):
    try:
        return crud.create_common_input(db, payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.put(
    "/project/{project_identifier}/common-input/{item_id}",
    response_model=schemas.CommonInputItem,
    tags=["Project Data"],
)
def update_project_common_input(
    project_identifier: str,
    item_id: int,
    payload: schemas.CommonInputUpdate,
    db: Session = Depends(get_project_db_session),
):
    updated = crud.update_common_input(db, item_id, payload)
    if not updated:
        raise HTTPException(status_code=404, detail="CommonInput item not found")
    return updated


@router.delete(
    "/project/{project_identifier}/common-input/{item_id}",
    tags=["Project Data"],
)
def delete_project_common_input(
    project_identifier: str,
    item_id: int,
    db: Session = Depends(get_project_db_session),
):
    deleted = crud.delete_common_input(db, item_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="CommonInput item not found")
    return {"message": "deleted", "id": item_id}


@router.get(
    "/project/{project_identifier}/calc-dictionary",
    response_model=List[schemas.CalcDictionaryEntry],
    tags=["Project Data"],
)
def read_project_calc_dictionary(
    project_identifier: str, db: Session = Depends(get_project_db_session)
):
    return crud.list_all_calc_dictionary_entries(db)


@router.post(
    "/project/{project_identifier}/calc-dictionary",
    response_model=schemas.CalcDictionaryEntry,
    tags=["Project Data"],
)
def create_project_calc_dictionary_entry(
    project_identifier: str,
    payload: schemas.ProjectCalcDictionaryEntryCreate,
    db: Session = Depends(get_project_db_session),
):
    family_id = payload.family_list_id
    if family_id is not None:
        family_item = crud.get_family_item(db, family_id)
        if not family_item:
            raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.create_project_calc_dictionary_entry(db, payload)


@router.patch(
    "/project/{project_identifier}/calc-dictionary/{entry_id}",
    response_model=schemas.CalcDictionaryEntry,
    tags=["Project Data"],
)
def update_project_calc_dictionary_entry(
    project_identifier: str,
    entry_id: int,
    payload: schemas.CalcDictionaryEntryUpdate,
    db: Session = Depends(get_project_db_session),
):
    entry = crud.get_calc_dictionary_entry(db, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    updates = payload.model_dump(exclude_none=True)
    if not updates:
        return entry
    updated = crud.update_calc_dictionary_entry(db, entry_id=entry_id, updates=updates)
    if not updated:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    return updated


@router.delete(
    "/project/{project_identifier}/calc-dictionary/{entry_id}",
    tags=["Project Data"],
)
def delete_project_calc_dictionary_entry(
    project_identifier: str,
    entry_id: int,
    db: Session = Depends(get_project_db_session),
):
    entry = crud.get_calc_dictionary_entry(db, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    updated = crud.update_calc_dictionary_entry(
        db, entry_id=entry_id, updates={"is_deleted": 1}
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    return {"message": "deleted", "id": entry_id}


@router.post(
    "/project/{project_identifier}/calc-dictionary/sync-with-common-input",
    response_model=schemas.CalcDictionarySyncResult,
    tags=["Project Data"],
)
def sync_project_calc_dictionary_with_common_input(
    project_identifier: str, db: Session = Depends(get_project_db_session)
):
    updated_entries = crud.sync_calc_dictionary_with_common_inputs(db)
    return schemas.CalcDictionarySyncResult(updated_entries=updated_entries)


@router.get(
    "/family-list/{item_id}/assignments",
    response_model=List[schemas.GwmFamilyAssign],
    tags=["Family List"],
)
def read_family_assignments(item_id: int, db: Session = Depends(get_db)):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.list_gwm_family_assignments(db, family_id=item_id)


@router.post(
    "/family-list/{item_id}/assignments",
    response_model=List[schemas.GwmFamilyAssign],
    tags=["Family List"],
)
def replace_family_assignments(
    item_id: int,
    payload: schemas.GwmFamilyAssignmentPayload,
    db: Session = Depends(get_db),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.replace_gwm_family_assignments(
        db, family_id=item_id, standard_item_ids=payload.standard_item_ids
    )


@router.patch(
    "/family-list/{item_id}/assignments/{assignment_id}",
    response_model=schemas.GwmFamilyAssign,
    tags=["Family List"],
)
def update_family_assignment_metadata(
    item_id: int,
    assignment_id: int,
    payload: schemas.GwmFamilyAssignUpdate,
    db: Session = Depends(get_db),
):
    updated = crud.update_gwm_family_assignment(
        db,
        family_id=item_id,
        assignment_id=assignment_id,
        updates=payload.model_dump(exclude_unset=True),
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Assignment not found")
    return updated


@router.post(
    "/family-list/{item_id}/assignments/{standard_item_id}",
    response_model=schemas.GwmFamilyAssign,
    tags=["Family List"],
)
def create_family_assignment(
    item_id: int,
    standard_item_id: int,
    db: Session = Depends(get_db),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    assignment = crud.create_gwm_family_assignment(
        db, family_id=item_id, standard_item_id=standard_item_id
    )
    return assignment


@router.get(
    "/project/{project_identifier}/family-list/",
    response_model=List[schemas.FamilyListItem],
    tags=["Project Data"],
)
def read_project_family_list(
    project_identifier: str, db: Session = Depends(get_project_db_session)
):
    return crud.list_family_items(db)


@router.post(
    "/project/{project_identifier}/family-list/",
    response_model=schemas.FamilyListItem,
    tags=["Project Data"],
)
def create_project_family_list_item(
    project_identifier: str,
    item: schemas.FamilyListCreate,
    db: Session = Depends(get_project_db_session),
):
    return crud.create_family_item(db, item)


@router.put(
    "/project/{project_identifier}/family-list/{item_id}",
    response_model=schemas.FamilyListItem,
    tags=["Project Data"],
)
def update_project_family_list_item(
    project_identifier: str,
    item_id: int,
    payload: schemas.FamilyListUpdate,
    db: Session = Depends(get_project_db_session),
):
    updated = crud.update_family_item(db, item_id, payload)
    if not updated:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return updated


@router.delete(
    "/project/{project_identifier}/family-list/{item_id}", tags=["Project Data"]
)
def delete_project_family_list_item(
    project_identifier: str, item_id: int, db: Session = Depends(get_project_db_session)
):
    deleted = crud.delete_family_item(db, item_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return {"message": "deleted", "id": item_id}


@router.get(
    "/project/{project_identifier}/family-list/{item_id}/calc-dictionary",
    response_model=List[schemas.CalcDictionaryEntry],
    tags=["Project Data"],
)
def read_project_family_calc_dictionary(
    project_identifier: str,
    item_id: int,
    db: Session = Depends(get_project_db_session),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.list_calc_dictionary_entries(db, family_item_id=item_id)


@router.post(
    "/project/{project_identifier}/family-list/{item_id}/calc-dictionary",
    response_model=schemas.CalcDictionaryEntry,
    tags=["Project Data"],
)
def create_project_family_calc_dictionary_entry(
    project_identifier: str,
    item_id: int,
    entry: schemas.CalcDictionaryEntryCreate,
    db: Session = Depends(get_project_db_session),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.create_calc_dictionary_entry(db, family_item_id=item_id, entry_in=entry)


@router.patch(
    "/project/{project_identifier}/family-list/{item_id}/calc-dictionary/{entry_id}",
    response_model=schemas.CalcDictionaryEntry,
    tags=["Project Data"],
)
def update_project_family_calc_dictionary_entry(
    project_identifier: str,
    item_id: int,
    entry_id: int,
    payload: schemas.CalcDictionaryEntryUpdate,
    db: Session = Depends(get_project_db_session),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    entry = crud.get_calc_dictionary_entry(db, entry_id)
    if not entry or entry.family_list_id != item_id:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    updates = payload.model_dump(exclude_none=True)
    if not updates:
        return entry
    updated = crud.update_calc_dictionary_entry(db, entry_id=entry_id, updates=updates)
    if not updated:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    return updated


@router.delete(
    "/project/{project_identifier}/family-list/{item_id}/calc-dictionary/{entry_id}",
    tags=["Project Data"],
)
def delete_project_family_calc_dictionary_entry(
    project_identifier: str,
    item_id: int,
    entry_id: int,
    db: Session = Depends(get_project_db_session),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    entry = crud.get_calc_dictionary_entry(db, entry_id)
    if not entry or entry.family_list_id != item_id:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    deleted = crud.delete_calc_dictionary_entry(db, entry_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Calc dictionary entry not found")
    return {"message": "deleted", "id": entry_id}


@router.get(
    "/project/{project_identifier}/family-list/{item_id}/revit-types",
    response_model=List[schemas.FamilyRevitType],
    tags=["Project Data"],
)
def read_project_family_revit_types(
    project_identifier: str,
    item_id: int,
    db: Session = Depends(get_project_db_session),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.list_family_revit_types(db, family_item_id=item_id)


@router.post(
    "/project/{project_identifier}/family-list/{item_id}/revit-types",
    response_model=List[schemas.FamilyRevitType],
    tags=["Project Data"],
)
def replace_project_family_revit_types(
    project_identifier: str,
    item_id: int,
    payload: schemas.FamilyRevitTypeListPayload,
    db: Session = Depends(get_project_db_session),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    entries = payload.entries
    if not entries:
        entries = [{"type_name": name} for name in payload.type_names]
    return crud.replace_family_revit_types(db, family_item_id=item_id, entries=entries)


@router.get(
    "/project/{project_identifier}/family-list/{item_id}/assignments",
    response_model=List[schemas.GwmFamilyAssign],
    tags=["Project Data"],
)
def read_project_family_assignments(
    project_identifier: str, item_id: int, db: Session = Depends(get_project_db_session)
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.list_gwm_family_assignments(db, family_id=item_id)


@router.post(
    "/project/{project_identifier}/family-list/{item_id}/assignments",
    response_model=List[schemas.GwmFamilyAssign],
    tags=["Project Data"],
)
def replace_project_family_assignments(
    project_identifier: str,
    item_id: int,
    payload: schemas.GwmFamilyAssignmentPayload,
    db: Session = Depends(get_project_db_session),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    return crud.replace_gwm_family_assignments(
        db, family_id=item_id, standard_item_ids=payload.standard_item_ids
    )


@router.patch(
    "/project/{project_identifier}/family-list/{item_id}/assignments/{assignment_id}",
    response_model=schemas.GwmFamilyAssign,
    tags=["Project Data"],
)
def update_project_family_assignment_metadata(
    project_identifier: str,
    item_id: int,
    assignment_id: int,
    payload: schemas.GwmFamilyAssignUpdate,
    db: Session = Depends(get_project_db_session),
):
    updated = crud.update_gwm_family_assignment(
        db,
        family_id=item_id,
        assignment_id=assignment_id,
        updates=payload.model_dump(exclude_unset=True),
    )
    if not updated:
        raise HTTPException(status_code=404, detail="Assignment not found")
    return updated


@router.post(
    "/project/{project_identifier}/family-list/{item_id}/assignments/{standard_item_id}",
    response_model=schemas.GwmFamilyAssign,
    tags=["Project Data"],
)
def create_project_family_assignment(
    project_identifier: str,
    item_id: int,
    standard_item_id: int,
    db: Session = Depends(get_project_db_session),
):
    family_item = crud.get_family_item(db, item_id)
    if not family_item:
        raise HTTPException(status_code=404, detail="FamilyList item not found")
    assignment = crud.create_gwm_family_assignment(
        db, family_id=item_id, standard_item_id=standard_item_id
    )
    return assignment


@router.get(
    "/project-db/",
    response_model=List[schemas.ProjectDbItem],
    tags=["Project DB"],
)
def list_project_databases():
    return project_db.list_project_dbs()


@router.get(
    "/project-db/backups/",
    response_model=List[schemas.ProjectDbBackupItem],
    tags=["Project DB"],
)
def list_project_database_backups():
    return project_db.list_project_db_backups()


@router.post(
    "/project-db/backups/{backup_file_name}/promote",
    response_model=schemas.ProjectDbItem,
    tags=["Project DB"],
)
def promote_project_database_backup(backup_file_name: str):
    try:
        return project_db.promote_backup_to_project_db(backup_file_name)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post(
    "/project-db/",
    response_model=schemas.ProjectDbItem,
    tags=["Project DB"],
)
def create_project_database(payload: schemas.ProjectDbCreate):
    try:
        return project_db.create_project_db(payload.display_name)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post(
    "/project-db/{file_name}/copy",
    response_model=schemas.ProjectDbItem,
    tags=["Project DB"],
)
def copy_project_database(file_name: str, payload: schemas.ProjectDbCopy):
    try:
        return project_db.copy_project_db(file_name, payload.display_name)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post(
    "/project-db/{file_name}/rename",
    response_model=schemas.ProjectDbItem,
    tags=["Project DB"],
)
def rename_project_database(file_name: str, payload: schemas.ProjectDbRename):
    try:
        return project_db.rename_project_db(file_name, payload.new_display_name)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post(
    "/project-db/{file_name}/backup",
    response_model=schemas.ProjectDbBackupResponse,
    tags=["Project DB"],
)
def backup_project_database(file_name: str):
    try:
        return project_db.backup_project_db(file_name)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.delete("/project-db/{file_name}", tags=["Project DB"])
def delete_project_database(file_name: str, admin_key: str):
    if admin_key != project_db.ADMIN_KEY:
        raise HTTPException(
            status_code=403, detail="Admin key is required for deletion"
        )
    try:
        project_db.delete_project_db(file_name)
        return {"message": "deleted", "file_name": file_name}
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))


@router.get(
    "/common-input/",
    response_model=List[schemas.CommonInputItem],
    tags=["Common Input"],
)
def list_common_input(db: Session = Depends(get_db)):
    return crud.list_common_inputs(db)


@router.post(
    "/common-input/",
    response_model=schemas.CommonInputItem,
    tags=["Common Input"],
)
def create_common_input(
    payload: schemas.CommonInputCreate, db: Session = Depends(get_db)
):
    try:
        return crud.create_common_input(db, payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.put(
    "/common-input/{item_id}",
    response_model=schemas.CommonInputItem,
    tags=["Common Input"],
)
def update_common_input(
    item_id: int, payload: schemas.CommonInputUpdate, db: Session = Depends(get_db)
):
    updated = crud.update_common_input(db, item_id, payload)
    if not updated:
        raise HTTPException(status_code=404, detail="CommonInput item not found")
    return updated


@router.delete("/common-input/{item_id}", tags=["Common Input"])
def delete_common_input(item_id: int, db: Session = Depends(get_db)):
    deleted = crud.delete_common_input(db, item_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="CommonInput item not found")
    return {"message": "deleted", "id": item_id}


@router.options("/{path:path}")
def cors_options(path: str):
    return Response(status_code=204)
